"""
CAO Salary Extraction - Truncated Files Descriptives

This script analyzes files where salary extraction failed due to truncation (files too large).
It creates a descriptive Excel report with information about these files.

USAGE:
    python scripts/excel_analysis/descriptives_truncated_files.py

INPUT:
    - Truncated files in: performance_logs/llm_analysis/max_tokens_truncated_4/
    - CAO info from: inputs/pdfs/extracted_cao_info.csv
    - Performance logs: performance_logs/llm_analysis/analysis_performance_salary.jsonl

OUTPUT:
    - Excel file: outputs/analysis/truncated_files_descriptives.xlsx
"""

# =============================================================================
# IMPORTS
# =============================================================================
import os
import sys
import json
import re
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple
from collections import defaultdict

# Add the parent directory to Python path
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font

# =============================================================================
# CONSTANTS
# =============================================================================
TRUNCATED_FOLDER_4 = "performance_logs/llm_analysis/max_tokens_truncated_4"
TRUNCATED_FOLDER_3 = "performance_logs/llm_analysis/max_tokens_truncated_3"
TRUNCATED_FOLDER_2 = "performance_logs/llm_analysis/max_tokens_truncated_2"
TRUNCATED_FOLDER_1 = "performance_logs/llm_analysis/max_tokens_truncated"
CAO_INFO_CSV = "inputs/pdfs/extracted_cao_info.csv"
# Fallback paths used when CAO_INFO_CSV is missing (e.g. project uses input_pdfs subfolder)
CAO_INFO_CSV_CANDIDATES = [
    "inputs/pdfs/extracted_cao_info.csv",
    "inputs/pdfs/input_pdfs/extracted_cao_info.csv",
]
SALARY_CSV = "outputs/excel/new_results/extracted_data_salary.csv"
PERFORMANCE_LOG_JSONL = "performance_logs/llm_analysis/analysis_performance_salary.jsonl"
LLM_ANALYSIS_FOLDER = "outputs/llm_analysis"
OUTPUT_EXCEL = "outputs/analysis/truncated_files_descriptives.xlsx"


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def extract_cao_number_and_filename(truncated_filename: str) -> tuple:
    """
    Extract CAO number and original filename from truncated file name.
    
    Format: {cao_number}_{filename}_truncated.txt
    
    Args:
        truncated_filename: Name of the truncated file
        
    Returns:
        tuple: (cao_number, original_filename) or (None, None) if parsing fails
    """
    # Remove .txt extension
    name_without_ext = truncated_filename.replace('_truncated.txt', '').replace('.txt', '')
    
    # Pattern: number at the start, followed by underscore, then the rest
    match = re.match(r'^(\d+)_(.+)$', name_without_ext)
    if match:
        cao_number = match.group(1)
        original_filename = match.group(2)
        return cao_number, original_filename
    
    return None, None


def extract_clean_filename(filename: str) -> str:
    """
    Extract a clean filename from a log filename (mirrors p4 extract_clean_filename for lookup).
    Removes _extract.json/.json, trailing .docx/.pdf/.doc, normalizes underscores, caps length.
    
    Args:
        filename: Original filename (e.g. "CAO_GHZ_2019-2021_definitief.docx_extract.json")
        
    Returns:
        Clean filename (e.g. "CAO_GHZ_2019-2021_definitief")
    """
    clean_name = filename
    if clean_name.endswith('_extract.json'):
        clean_name = clean_name[:-13]
    elif clean_name.endswith('.json'):
        clean_name = clean_name[:-5]
    for ext in ['.docx', '.pdf', '.doc']:
        if clean_name.endswith(ext):
            clean_name = clean_name[:-len(ext)]
            break
    clean_name = re.sub(r'_+', '_', clean_name)
    clean_name = clean_name.strip('_')
    if len(clean_name) > 100:
        clean_name = clean_name[:100].rstrip('_')
    return clean_name


def _first_column(df: pd.DataFrame, alternatives: List[str]) -> Optional[str]:
    """Return the first column name that exists in df, or None."""
    for name in alternatives:
        if name in df.columns:
            return name
    return None


def load_cao_info(cao_info_path: str) -> Dict[str, Dict]:
    """
    Load CAO information from CSV file.
    Accepts alternate column names (e.g. cao_numb, kennisgeving) for compatibility.
    
    Args:
        cao_info_path: Path to CAO info CSV (config-driven path may differ; see project config)
        
    Returns:
        dict: Mapping from (cao_number, pdf_name) to CAO metadata
    """
    cao_info = {}
    
    # Use first existing path: requested path, then candidate fallbacks
    effective_path = cao_info_path
    if not os.path.exists(effective_path):
        for candidate in CAO_INFO_CSV_CANDIDATES:
            if os.path.exists(candidate):
                effective_path = candidate
                print(f"  Using CAO info from: {effective_path}")
                break
    if not os.path.exists(effective_path):
        print(f"  Warning: CAO info file not found (tried: {cao_info_path} and {len(CAO_INFO_CSV_CANDIDATES)} candidates)")
        return cao_info
    
    try:
        df = pd.read_csv(effective_path, sep=';', encoding='utf-8')
        print(f"  Loaded {len(df)} CAO info records")
        
        col_cao = _first_column(df, ['cao_number', 'cao_numb'])
        col_pdf = _first_column(df, ['pdf_name', 'original_filenam', 'original_filename'])
        col_id = _first_column(df, ['id'])
        col_ingang = _first_column(df, ['ingangsdatum'])
        col_expir = _first_column(df, ['expiratiedatum'])
        col_kennis = _first_column(df, ['datum_kennisgeving', 'kennisgeving'])
        col_sbi = _first_column(df, ['sbi_code'])
        col_sector = _first_column(df, ['sector'])
        
        if not col_cao or not col_pdf:
            print(f"  Warning: CAO info CSV missing required columns (cao_number, pdf_name)")
            return cao_info
        
        for _, row in df.iterrows():
            cao_number = str(row.get(col_cao, '')) if col_cao else ''
            pdf_name = str(row.get(col_pdf, '')) if col_pdf else ''
            if pd.isna(cao_number) or pd.isna(pdf_name):
                continue
            cao_number = cao_number.strip()
            pdf_name = str(pdf_name).strip()
            if not cao_number or not pdf_name:
                continue
            key = (cao_number, pdf_name)
            def _val(col): return ('' if pd.isna(row.get(col)) else str(row.get(col))) if col else ''
            cao_info[key] = {
                'cao_number': cao_number,
                'pdf_name': pdf_name,
                'id': _val(col_id),
                'ingangsdatum': _val(col_ingang),
                'expiratiedatum': _val(col_expir),
                'datum_kennisgeving': _val(col_kennis),
                'sbi_code': _val(col_sbi),
                'sector': _val(col_sector)
            }
    except Exception as e:
        print(f"  Warning: Could not load CAO info: {e}")
    
    return cao_info


def match_cao_info(cao_number: str, filename: str, cao_info_dict: Dict) -> Optional[Dict]:
    """
    Match filename to CAO info using fuzzy matching.
    Tries exact key (cao_number, filename + '.pdf') first when CSV stores pdf_name with .pdf.
    
    Args:
        cao_number: CAO number
        filename: Original filename (clean, without .pdf)
        cao_info_dict: CAO info dictionary
        
    Returns:
        CAO info dict if found, None otherwise
    """
    # Try exact key as in p5: expected_pdf_name = filename + '.pdf'
    expected_pdf_name = filename + '.pdf'
    if (cao_number, expected_pdf_name) in cao_info_dict:
        return cao_info_dict[(cao_number, expected_pdf_name)]
    
    # Normalize filename for matching (remove extensions, spaces, etc.)
    def normalize(s: str) -> str:
        return s.replace(' ', '').replace('-', '').replace('_', '').lower()
    
    normalized_filename = normalize(filename)
    
    # Try fuzzy match: normalized substring
    for (cao_num, pdf_name), info in cao_info_dict.items():
        if cao_num == cao_number:
            normalized_pdf = normalize(pdf_name)
            if normalized_pdf in normalized_filename or normalized_filename in normalized_pdf:
                return info
    
    # Try by CAO number only
    for (cao_num, pdf_name), info in cao_info_dict.items():
        if cao_num == cao_number:
            return info  # Return first match for this CAO number
    
    return None


def load_performance_logs(jsonl_path: str) -> Tuple[Dict[str, Dict], Dict[Tuple[str, str], Dict]]:
    """
    Load performance logs to get file size and other metadata.
    
    Args:
        jsonl_path: Path to performance log JSONL file
        
    Returns:
        Tuple of (logs_by_filename, logs_by_cao_clean). logs_by_filename maps full filename to entry;
        logs_by_cao_clean maps (cao_number, clean_filename) to entry for lookup from truncated file names.
    """
    logs = {}
    logs_by_cao_clean = {}
    
    if not os.path.exists(jsonl_path):
        print(f"  Warning: Performance log file not found: {jsonl_path}")
        return logs, logs_by_cao_clean
    
    try:
        with open(jsonl_path, 'r', encoding='utf-8') as f:
            for line in f:
                if line.strip():
                    try:
                        entry = json.loads(line)
                        filename = entry.get('filename', '')
                        if filename:
                            logs[filename] = entry
                            cao_number = str(entry.get('cao_number', ''))
                            clean = extract_clean_filename(filename)
                            if cao_number and clean:
                                logs_by_cao_clean[(cao_number, clean)] = entry
                    except json.JSONDecodeError:
                        continue
        
        print(f"  Loaded {len(logs)} performance log entries")
    except Exception as e:
        print(f"  Warning: Could not load performance logs: {e}")
    
    return logs, logs_by_cao_clean


def get_perf_log(original_filename: str, cao_number: str,
                 performance_logs: Dict[str, Dict],
                 performance_logs_by_cao_clean: Dict[Tuple[str, str], Dict]) -> Dict:
    """
    Resolve performance log entry for a truncated file by trying multiple keys.
    
    Args:
        original_filename: Clean filename from truncated file name
        cao_number: CAO number
        performance_logs: Dict keyed by full log filename
        performance_logs_by_cao_clean: Dict keyed by (cao_number, clean_filename)
        
    Returns:
        Log entry dict or empty dict if not found
    """
    perf = performance_logs.get(original_filename, {})
    if perf:
        return perf
    perf = performance_logs.get(original_filename + '_extract.json', {})
    if perf:
        return perf
    return performance_logs_by_cao_clean.get((cao_number, original_filename), {})


def get_file_size_mb(file_path: Path) -> float:
    """Get file size in MB."""
    try:
        return file_path.stat().st_size / (1024 * 1024)
    except:
        return 0.0


# =============================================================================
# DATA COLLECTION
# =============================================================================

def collect_truncated_files(truncated_folder: str, truncation_level: str = "4") -> List[Dict]:
    """
    Collect information about all truncated files in a folder.
    
    Args:
        truncated_folder: Path to truncated folder
        truncation_level: Level of truncation ("1", "2", "3", or "4")
        
    Returns:
        List of dictionaries with file information
    """
    truncated_files = []
    folder_path = Path(truncated_folder)
    
    if not folder_path.exists():
        return truncated_files
    
    for truncated_file in folder_path.glob("*.txt"):
        cao_number, original_filename = extract_cao_number_and_filename(truncated_file.name)
        
        if cao_number and original_filename:
            file_size_mb = get_file_size_mb(truncated_file)
            truncated_files.append({
                'truncated_filename': truncated_file.name,
                'cao_number': cao_number,
                'original_filename': original_filename,
                'file_size_mb': file_size_mb,
                'file_path': str(truncated_file),
                'truncation_level': truncation_level
            })
    
    return truncated_files


def load_successful_files(salary_csv_path: str) -> Set[tuple]:
    """
    Load successfully extracted files from salary CSV.
    
    Args:
        salary_csv_path: Path to salary CSV
        
    Returns:
        Set of (cao_number, filename) tuples
    """
    successful = set()
    
    if not os.path.exists(salary_csv_path):
        print(f"  Warning: Salary CSV not found: {salary_csv_path}")
        return successful
    
    try:
        df = pd.read_csv(salary_csv_path, sep=';', encoding='utf-8', nrows=1000000)  # Limit to avoid memory issues
        if 'cao_number' in df.columns and 'file_name' in df.columns:
            # Get unique combinations of cao_number and file_name
            unique_files = df[['cao_number', 'file_name']].drop_duplicates()
            for _, row in unique_files.iterrows():
                cao_number = str(row.get('cao_number', ''))
                filename = str(row.get('file_name', ''))
                if cao_number and filename:
                    # Normalize filename (remove extensions)
                    filename_base = filename.replace('_extract.json', '').replace('.json', '').replace('.pdf', '')
                    successful.add((cao_number, filename_base))
        print(f"  Loaded {len(successful)} unique successful files from salary CSV")
    except Exception as e:
        print(f"  Warning: Could not load salary CSV: {e}")
    
    return successful


# =============================================================================
# SHEET GENERATION
# =============================================================================

def create_sheet_00_overview(truncated_files: List[Dict], cao_info_dict: Dict,
                            performance_logs: Dict,
                            performance_logs_by_cao_clean: Dict[Tuple[str, str], Dict]) -> pd.DataFrame:
    """
    Create overview sheet with all truncated files and their metadata.
    
    Args:
        truncated_files: List of truncated file info
        cao_info_dict: CAO info dictionary
        performance_logs: Performance log entries keyed by filename
        performance_logs_by_cao_clean: Performance log entries keyed by (cao_number, clean_filename)
        
    Returns:
        DataFrame with overview
    """
    rows = []
    
    for file_info in truncated_files:
        cao_number = file_info['cao_number']
        original_filename = file_info['original_filename']
        
        # Match CAO info
        cao_info = match_cao_info(cao_number, original_filename, cao_info_dict)
        
        # Get performance log info (try multiple keys to match log entries)
        perf_log = get_perf_log(original_filename, cao_number, performance_logs, performance_logs_by_cao_clean)
        
        row = {
            'cao_number': cao_number,
            'original_filename': original_filename,
            'truncated_file_size_mb': file_info['file_size_mb'],
            'id': cao_info.get('id', '') if cao_info else '',
            'pdf_name': cao_info.get('pdf_name', '') if cao_info else '',
            'ingangsdatum': cao_info.get('ingangsdatum', '') if cao_info else '',
            'expiratiedatum': cao_info.get('expiratiedatum', '') if cao_info else '',
            'datum_kennisgeving': cao_info.get('datum_kennisgeving', '') if cao_info else '',
            'sbi_code': cao_info.get('sbi_code', '') if cao_info else '',
            'sector': cao_info.get('sector', '') if cao_info else '',
            'log_file_size_mb': perf_log.get('file_size_mb', np.nan),
            'log_processing_time': perf_log.get('processing_time_seconds', np.nan),
            'log_input_tokens': perf_log.get('input_tokens', np.nan),
            'log_output_tokens': perf_log.get('output_tokens', np.nan),
            'log_total_tokens': perf_log.get('total_tokens', np.nan),
            'log_error_message': perf_log.get('error_message', '')
        }
        rows.append(row)
    
    df = pd.DataFrame(rows)
    return df


def create_sheet_01_by_cao(truncated_files: List[Dict], cao_info_dict: Dict) -> pd.DataFrame:
    """
    Create summary by CAO number.
    
    Args:
        truncated_files: List of truncated file info
        cao_info_dict: CAO info dictionary
        
    Returns:
        DataFrame with summary by CAO
    """
    cao_stats = defaultdict(lambda: {
        'n_files': 0,
        'cao_numbers': set(),
        'sectors': set(),
        'sbi_codes': set(),
        'filenames': []
    })
    
    for file_info in truncated_files:
        cao_number = file_info['cao_number']
        original_filename = file_info['original_filename']
        
        cao_info = match_cao_info(cao_number, original_filename, cao_info_dict)
        
        cao_stats[cao_number]['n_files'] += 1
        cao_stats[cao_number]['cao_numbers'].add(cao_number)
        cao_stats[cao_number]['filenames'].append(original_filename)
        
        if cao_info:
            if cao_info.get('sector'):
                cao_stats[cao_number]['sectors'].add(cao_info['sector'])
            if cao_info.get('sbi_code'):
                cao_stats[cao_number]['sbi_codes'].add(cao_info['sbi_code'])
    
    rows = []
    for cao_number, stats in sorted(cao_stats.items()):
        rows.append({
            'cao_number': cao_number,
            'n_truncated_files': stats['n_files'],
            'sectors': ', '.join(sorted(stats['sectors'])) if stats['sectors'] else '',
            'sbi_codes': ', '.join(sorted(stats['sbi_codes'])) if stats['sbi_codes'] else '',
            'filenames': ' | '.join(stats['filenames'][:3]) + (' ...' if len(stats['filenames']) > 3 else '')
        })
    
    df = pd.DataFrame(rows)
    return df


def create_sheet_02_by_sector(truncated_files: List[Dict], cao_info_dict: Dict) -> pd.DataFrame:
    """
    Create summary by sector.
    
    Args:
        truncated_files: List of truncated file info
        cao_info_dict: CAO info dictionary
        
    Returns:
        DataFrame with summary by sector
    """
    sector_stats = defaultdict(lambda: {
        'n_files': 0,
        'n_caos': set(),
        'cao_numbers': []
    })
    
    for file_info in truncated_files:
        cao_number = file_info['cao_number']
        original_filename = file_info['original_filename']
        
        cao_info = match_cao_info(cao_number, original_filename, cao_info_dict)
        sector = cao_info.get('sector', 'Unknown') if cao_info else 'Unknown'
        
        sector_stats[sector]['n_files'] += 1
        sector_stats[sector]['n_caos'].add(cao_number)
        sector_stats[sector]['cao_numbers'].append(cao_number)
    
    rows = []
    for sector, stats in sorted(sector_stats.items()):
        rows.append({
            'sector': sector,
            'n_truncated_files': stats['n_files'],
            'n_unique_caos': len(stats['n_caos']),
            'cao_numbers': ', '.join(sorted(stats['n_caos']))
        })
    
    df = pd.DataFrame(rows)
    return df


def create_sheet_03_statistics(truncated_files: List[Dict], performance_logs: Dict,
                               performance_logs_by_cao_clean: Dict[Tuple[str, str], Dict]) -> pd.DataFrame:
    """
    Create overall statistics.
    
    Args:
        truncated_files: List of truncated file info
        performance_logs: Performance log entries keyed by filename
        performance_logs_by_cao_clean: Performance log entries keyed by (cao_number, clean_filename)
        
    Returns:
        DataFrame with statistics
    """
    stats = []
    
    # Basic counts
    stats.append({'metric': 'Total truncated files', 'value': len(truncated_files)})
    
    unique_caos = set(f['cao_number'] for f in truncated_files)
    stats.append({'metric': 'Unique CAO numbers', 'value': len(unique_caos)})
    
    # File sizes
    file_sizes = [f['file_size_mb'] for f in truncated_files if f['file_size_mb'] > 0]
    if file_sizes:
        stats.append({'metric': 'Mean truncated file size (MB)', 'value': np.mean(file_sizes)})
        stats.append({'metric': 'Median truncated file size (MB)', 'value': np.median(file_sizes)})
        stats.append({'metric': 'Min truncated file size (MB)', 'value': np.min(file_sizes)})
        stats.append({'metric': 'Max truncated file size (MB)', 'value': np.max(file_sizes)})
    
    # Performance log statistics
    log_sizes = []
    log_times = []
    log_input_tokens = []
    log_output_tokens = []
    
    for file_info in truncated_files:
        original_filename = file_info['original_filename']
        cao_number = file_info['cao_number']
        perf_log = get_perf_log(original_filename, cao_number, performance_logs, performance_logs_by_cao_clean)
        
        if perf_log.get('file_size_mb'):
            log_sizes.append(perf_log['file_size_mb'])
        if perf_log.get('processing_time_seconds'):
            log_times.append(perf_log['processing_time_seconds'])
        if perf_log.get('input_tokens'):
            log_input_tokens.append(perf_log['input_tokens'])
        if perf_log.get('output_tokens'):
            log_output_tokens.append(perf_log['output_tokens'])
    
    if log_sizes:
        stats.append({'metric': 'Mean original file size (MB)', 'value': np.mean(log_sizes)})
        stats.append({'metric': 'Median original file size (MB)', 'value': np.median(log_sizes)})
    
    if log_times:
        stats.append({'metric': 'Mean processing time (seconds)', 'value': np.mean(log_times)})
        stats.append({'metric': 'Median processing time (seconds)', 'value': np.median(log_times)})
    
    if log_input_tokens:
        stats.append({'metric': 'Mean input tokens', 'value': np.mean(log_input_tokens)})
        stats.append({'metric': 'Median input tokens', 'value': np.median(log_input_tokens)})
    
    if log_output_tokens:
        stats.append({'metric': 'Mean output tokens', 'value': np.mean(log_output_tokens)})
        stats.append({'metric': 'Median output tokens', 'value': np.median(log_output_tokens)})
    
    df = pd.DataFrame(stats)
    return df


# =============================================================================
# MAIN EXECUTION
# =============================================================================

def check_if_truncated_file_was_saved(truncated_file_info: Dict, llm_analysis_folder: Path) -> bool:
    """
    Check if a truncated file was successfully saved after retry.
    
    Args:
        truncated_file_info: Info about truncated file
        llm_analysis_folder: Path to llm_analysis folder
        
    Returns:
        True if file exists in salary folder, False otherwise
    """
    cao_number = truncated_file_info['cao_number']
    original_filename = truncated_file_info['original_filename']
    # Normalize filename
    base_filename = original_filename.replace('_extract.json', '').replace('.json', '').replace('.pdf', '')
    
    salary_file = llm_analysis_folder / 'salary' / cao_number / f"{base_filename}_analysis.json"
    return salary_file.exists()


def check_if_file_exists_in_salary_folder(cao_number: str, filename: str, llm_analysis_folder: Path) -> bool:
    """
    Check if a file exists in the salary folder.
    
    Args:
        cao_number: CAO number
        filename: Base filename (normalized)
        llm_analysis_folder: Path to llm_analysis folder
        
    Returns:
        True if file exists, False otherwise
    """
    salary_file = llm_analysis_folder / 'salary' / cao_number / f"{filename}_analysis.json"
    return salary_file.exists()


def create_sheet_04_file_status_comparison(truncated_files_all: List[Dict], successful_files: Set[tuple],
                                           cao_info_dict: Dict, llm_analysis_folder: Path) -> pd.DataFrame:
    """
    Create comparison showing file status: successful, truncated (by level), or missing.
    
    Args:
        truncated_files_all: All truncated files from all levels
        successful_files: Set of (cao_number, filename) tuples that were successfully extracted
        cao_info_dict: CAO info dictionary
        
    Returns:
        DataFrame with file status comparison
    """
    # Organize truncated files by level
    truncated_by_level = defaultdict(list)
    for file_info in truncated_files_all:
        level = file_info.get('truncation_level', '4')
        truncated_by_level[level].append(file_info)
    
    # Get all unique files from CAO info
    all_expected_files = set()
    for (cao_number, pdf_name), info in cao_info_dict.items():
        # Normalize filename
        filename_base = pdf_name.replace('_extract.json', '').replace('.json', '').replace('.pdf', '')
        all_expected_files.add((cao_number, filename_base))
    
    # Categorize files
    rows = []
    
    # Files in truncated_4 (all attempts exhausted)
    truncated_4_set = set()
    for file_info in truncated_by_level.get('4', []):
        cao_number = file_info['cao_number']
        filename = file_info['original_filename'].replace('_extract.json', '').replace('.json', '').replace('.pdf', '')
        truncated_4_set.add((cao_number, filename))
        was_saved = check_if_truncated_file_was_saved(file_info, llm_analysis_folder)
        rows.append({
            'cao_number': cao_number,
            'filename': filename,
            'status': 'truncated_4_all_attempts_exhausted',
            'truncation_level': '4',
            'was_saved_after_retry': was_saved
        })
    
    # Files in truncated_3 (might be retried)
    truncated_3_set = set()
    for file_info in truncated_by_level.get('3', []):
        cao_number = file_info['cao_number']
        filename = file_info['original_filename'].replace('_extract.json', '').replace('.json', '').replace('.pdf', '')
        truncated_3_set.add((cao_number, filename))
        was_saved = check_if_truncated_file_was_saved(file_info, llm_analysis_folder)
        rows.append({
            'cao_number': cao_number,
            'filename': filename,
            'status': 'truncated_3_might_be_retried',
            'truncation_level': '3',
            'was_saved_after_retry': was_saved
        })
    
    # Files in truncated_2 (might be retried)
    truncated_2_set = set()
    for file_info in truncated_by_level.get('2', []):
        cao_number = file_info['cao_number']
        filename = file_info['original_filename'].replace('_extract.json', '').replace('.json', '').replace('.pdf', '')
        truncated_2_set.add((cao_number, filename))
        was_saved = check_if_truncated_file_was_saved(file_info, llm_analysis_folder)
        rows.append({
            'cao_number': cao_number,
            'filename': filename,
            'status': 'truncated_2_might_be_retried',
            'truncation_level': '2',
            'was_saved_after_retry': was_saved
        })
    
    # Files in truncated_1 (might be retried)
    truncated_1_set = set()
    for file_info in truncated_by_level.get('1', []):
        cao_number = file_info['cao_number']
        filename = file_info['original_filename'].replace('_extract.json', '').replace('.json', '').replace('.pdf', '')
        truncated_1_set.add((cao_number, filename))
        was_saved = check_if_truncated_file_was_saved(file_info, llm_analysis_folder)
        rows.append({
            'cao_number': cao_number,
            'filename': filename,
            'status': 'truncated_1_might_be_retried',
            'truncation_level': '1',
            'was_saved_after_retry': was_saved
        })
    
    # Successfully extracted files
    for cao_number, filename in successful_files:
        if (cao_number, filename) not in truncated_4_set and \
           (cao_number, filename) not in truncated_3_set and \
           (cao_number, filename) not in truncated_2_set and \
           (cao_number, filename) not in truncated_1_set:
            rows.append({
                'cao_number': cao_number,
                'filename': filename,
                'status': 'successfully_extracted',
                'truncation_level': '',
                'was_saved_after_retry': False  # Not applicable for successful files
            })
    
    # Files in CAO info but missing from salary CSV (regardless of truncated status)
    # This is the key comparison: expected files vs what's actually in the salary CSV
    missing_from_csv = all_expected_files - successful_files
    for cao_number, filename in missing_from_csv:
        # Check if file exists in salary folder (might have been extracted but not in CSV for some reason)
        exists_in_folder = check_if_file_exists_in_salary_folder(cao_number, filename, llm_analysis_folder)
        
        # Determine status based on whether it's in truncated folders
        if (cao_number, filename) in truncated_4_set:
            status = 'truncated_4_all_attempts_exhausted'
        elif (cao_number, filename) in truncated_3_set:
            status = 'truncated_3_might_be_retried'
        elif (cao_number, filename) in truncated_2_set:
            status = 'truncated_2_might_be_retried'
        elif (cao_number, filename) in truncated_1_set:
            status = 'truncated_1_might_be_retried'
        else:
            status = 'missing_from_salary_csv'
        
        # Only add if not already in rows (avoid duplicates with truncated files)
        if not any(r.get('cao_number') == cao_number and r.get('filename') == filename for r in rows):
            rows.append({
                'cao_number': cao_number,
                'filename': filename,
                'status': status,
                'truncation_level': '',
                'was_saved_after_retry': exists_in_folder
            })
    
    df = pd.DataFrame(rows)
    return df


def create_sheet_05_summary_counts(truncated_files_all: List[Dict], successful_files: Set[tuple],
                                  cao_info_dict: Dict, llm_analysis_folder: Path) -> pd.DataFrame:
    """
    Create summary counts by status.
    
    Args:
        truncated_files_all: All truncated files from all levels
        successful_files: Set of (cao_number, filename) tuples that were successfully extracted (from salary CSV)
        cao_info_dict: CAO info dictionary
        
    Returns:
        DataFrame with summary counts
    """
    # Count by truncation level and check which were saved
    truncated_by_level = defaultdict(int)
    truncated_saved_by_level = defaultdict(int)
    for file_info in truncated_files_all:
        level = file_info.get('truncation_level', '4')
        truncated_by_level[level] += 1
        if check_if_truncated_file_was_saved(file_info, llm_analysis_folder):
            truncated_saved_by_level[level] += 1
    
    # Get all expected files from CAO info
    all_expected_files = set()
    for (cao_number, pdf_name), info in cao_info_dict.items():
        filename_base = pdf_name.replace('_extract.json', '').replace('.json', '').replace('.pdf', '')
        all_expected_files.add((cao_number, filename_base))
    
    # Files that are in expected but not in successful (missing from salary CSV)
    missing_from_salary_csv = all_expected_files - successful_files
    
    # Count unique IDs
    unique_ids_successful = set()
    unique_ids_truncated_4 = set()
    unique_ids_truncated_3 = set()
    unique_ids_truncated_2 = set()
    unique_ids_truncated_1 = set()
    unique_ids_missing = set()
    
    for cao_number, filename in successful_files:
        cao_info = match_cao_info(cao_number, filename, cao_info_dict)
        if cao_info and cao_info.get('id'):
            unique_ids_successful.add(cao_info['id'])
    
    for file_info in truncated_files_all:
        cao_number = file_info['cao_number']
        filename = file_info['original_filename'].replace('_extract.json', '').replace('.json', '').replace('.pdf', '')
        cao_info = match_cao_info(cao_number, filename, cao_info_dict)
        if cao_info and cao_info.get('id'):
            level = file_info.get('truncation_level', '4')
            if level == '4':
                unique_ids_truncated_4.add(cao_info['id'])
            elif level == '3':
                unique_ids_truncated_3.add(cao_info['id'])
            elif level == '2':
                unique_ids_truncated_2.add(cao_info['id'])
            elif level == '1':
                unique_ids_truncated_1.add(cao_info['id'])
    
    all_processed = successful_files | set(
        (f['cao_number'], f['original_filename'].replace('_extract.json', '').replace('.json', '').replace('.pdf', ''))
        for f in truncated_files_all
    )
    missing_files = all_expected_files - all_processed
    for cao_number, filename in missing_files:
        cao_info = match_cao_info(cao_number, filename, cao_info_dict)
        if cao_info and cao_info.get('id'):
            unique_ids_missing.add(cao_info['id'])
    
    # Count unique IDs for missing files
    unique_ids_missing_from_csv = set()
    for cao_number, filename in missing_from_salary_csv:
        cao_info = match_cao_info(cao_number, filename, cao_info_dict)
        if cao_info and cao_info.get('id'):
            unique_ids_missing_from_csv.add(cao_info['id'])
    
    rows = [
        {'status': 'Successfully extracted (in salary CSV)', 'n_files': len(successful_files), 'n_unique_ids': len(unique_ids_successful), 'n_saved_after_retry': 0},
        {'status': 'Missing from salary CSV (expected but not extracted)', 'n_files': len(missing_from_salary_csv), 'n_unique_ids': len(unique_ids_missing_from_csv), 'n_saved_after_retry': 0},
        {'status': 'Truncated level 4 (all attempts exhausted)', 'n_files': truncated_by_level['4'], 'n_unique_ids': len(unique_ids_truncated_4), 'n_saved_after_retry': truncated_saved_by_level['4']},
        {'status': 'Truncated level 3 (might be retried)', 'n_files': truncated_by_level['3'], 'n_unique_ids': len(unique_ids_truncated_3), 'n_saved_after_retry': truncated_saved_by_level['3']},
        {'status': 'Truncated level 2 (might be retried)', 'n_files': truncated_by_level['2'], 'n_unique_ids': len(unique_ids_truncated_2), 'n_saved_after_retry': truncated_saved_by_level['2']},
        {'status': 'Truncated level 1 (might be retried)', 'n_files': truncated_by_level['1'], 'n_unique_ids': len(unique_ids_truncated_1), 'n_saved_after_retry': truncated_saved_by_level['1']},
        {'status': 'TOTAL (from CAO info)', 'n_files': len(all_expected_files), 'n_unique_ids': len(cao_info_dict), 'n_saved_after_retry': 0}
    ]
    
    df = pd.DataFrame(rows)
    return df


def main():
    """Main entry point."""
    print("="*80)
    print("CAO Truncated Files Descriptives")
    print("="*80)
    
    # Load data
    print("\nLoading data...")
    
    # Collect truncated files from all levels
    print("\nScanning truncated files...")
    truncated_files_4 = collect_truncated_files(TRUNCATED_FOLDER_4, "4")
    print(f"  Found {len(truncated_files_4)} files in truncated_4")
    truncated_files_3 = collect_truncated_files(TRUNCATED_FOLDER_3, "3")
    print(f"  Found {len(truncated_files_3)} files in truncated_3")
    truncated_files_2 = collect_truncated_files(TRUNCATED_FOLDER_2, "2")
    print(f"  Found {len(truncated_files_2)} files in truncated_2")
    truncated_files_1 = collect_truncated_files(TRUNCATED_FOLDER_1, "1")
    print(f"  Found {len(truncated_files_1)} files in truncated_1")
    
    truncated_files_all = truncated_files_4 + truncated_files_3 + truncated_files_2 + truncated_files_1
    print(f"  Total truncated files: {len(truncated_files_all)}")
    
    # For backward compatibility, use truncated_4 for existing sheets
    truncated_files = truncated_files_4
    
    print("\nLoading CAO info and performance logs...")
    cao_info_dict = load_cao_info(CAO_INFO_CSV)
    performance_logs, performance_logs_by_cao_clean = load_performance_logs(PERFORMANCE_LOG_JSONL)
    
    print("\nLoading successful files from salary CSV...")
    successful_files = load_successful_files(SALARY_CSV)
    
    # Create sheets
    print("\nGenerating sheets...")
    sheets = {}
    
    try:
        sheets["00_overview"] = create_sheet_00_overview(
            truncated_files, cao_info_dict, performance_logs, performance_logs_by_cao_clean)
        print(f"  Created sheet '00_overview' with {len(sheets['00_overview'])} rows")
        # Optional diagnostics: how many rows have CAO info vs performance log data
        if len(sheets["00_overview"]) > 0:
            n_cao_match = (sheets["00_overview"]["id"].astype(str).str.strip() != "").sum()
            n_perf_match = pd.notna(sheets["00_overview"].get("log_file_size_mb", pd.Series())).sum()
            print(f"  Overview: {n_cao_match} rows with CAO info match, {n_perf_match} rows with performance log match")
    except Exception as e:
        print(f"  Error creating overview sheet: {e}")
        import traceback
        traceback.print_exc()
        sheets["00_overview"] = pd.DataFrame()
    
    try:
        sheets["01_by_cao"] = create_sheet_01_by_cao(truncated_files, cao_info_dict)
        print(f"  Created sheet '01_by_cao' with {len(sheets['01_by_cao'])} rows")
    except Exception as e:
        print(f"  Error creating by_cao sheet: {e}")
        import traceback
        traceback.print_exc()
        sheets["01_by_cao"] = pd.DataFrame()
    
    try:
        sheets["02_by_sector"] = create_sheet_02_by_sector(truncated_files, cao_info_dict)
        print(f"  Created sheet '02_by_sector' with {len(sheets['02_by_sector'])} rows")
    except Exception as e:
        print(f"  Error creating by_sector sheet: {e}")
        import traceback
        traceback.print_exc()
        sheets["02_by_sector"] = pd.DataFrame()
    
    try:
        sheets["03_statistics"] = create_sheet_03_statistics(
            truncated_files, performance_logs, performance_logs_by_cao_clean)
        print(f"  Created sheet '03_statistics' with {len(sheets['03_statistics'])} rows")
    except Exception as e:
        print(f"  Error creating statistics sheet: {e}")
        import traceback
        traceback.print_exc()
        sheets["03_statistics"] = pd.DataFrame()
    
    try:
        llm_analysis_path = Path(LLM_ANALYSIS_FOLDER)
        sheets["04_file_status_comparison"] = create_sheet_04_file_status_comparison(
            truncated_files_all, successful_files, cao_info_dict, llm_analysis_path)
        print(f"  Created sheet '04_file_status_comparison' with {len(sheets['04_file_status_comparison'])} rows")
    except Exception as e:
        print(f"  Error creating file status comparison sheet: {e}")
        import traceback
        traceback.print_exc()
        sheets["04_file_status_comparison"] = pd.DataFrame()
    
    try:
        llm_analysis_path = Path(LLM_ANALYSIS_FOLDER)
        sheets["05_summary_counts"] = create_sheet_05_summary_counts(
            truncated_files_all, successful_files, cao_info_dict, llm_analysis_path)
        print(f"  Created sheet '05_summary_counts' with {len(sheets['05_summary_counts'])} rows")
    except Exception as e:
        print(f"  Error creating summary counts sheet: {e}")
        import traceback
        traceback.print_exc()
        sheets["05_summary_counts"] = pd.DataFrame()
    
    # Write to Excel
    print(f"\nWriting to Excel: {OUTPUT_EXCEL}")
    try:
        # Ensure output directory exists
        os.makedirs(os.path.dirname(OUTPUT_EXCEL), exist_ok=True)
        
        # First write all dataframes
        with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl') as writer:
            for sheet_name, sheet_df in sheets.items():
                if len(sheet_df) > 0:
                    sheet_df_clean = sheet_df.where(pd.notna(sheet_df), None)
                    for col in sheet_df_clean.columns:
                        if sheet_df_clean[col].dtype == 'object':
                            sheet_df_clean[col] = sheet_df_clean[col].fillna('')
                    sheet_df_clean.to_excel(writer, sheet_name=sheet_name, index=False)
                    print(f"  Wrote sheet '{sheet_name}' with {len(sheet_df_clean)} rows")
                else:
                    sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    print(f"  Wrote empty sheet '{sheet_name}'")
        
        # Add explanatory notes
        wb = load_workbook(OUTPUT_EXCEL)
        
        notes = {
            "00_overview": [
                "NOTES:",
                "This sheet lists all files where salary extraction failed due to truncation (files too large).",
                "",
                "cao_number: CAO number extracted from truncated filename.",
                "original_filename: Original filename before truncation.",
                "truncated_file_size_mb: Size of the truncated response file (MB).",
                "id, pdf_name, ingangsdatum, expiratiedatum, datum_kennisgeving: CAO metadata from CAO info CSV.",
                "sbi_code, sector: Sector information from CAO info CSV.",
                "log_file_size_mb: Original file size from performance logs (if available).",
                "log_processing_time: Processing time in seconds (if available).",
                "log_input_tokens, log_output_tokens, log_total_tokens: Token usage from performance logs (if available).",
                "log_error_message: Error message from performance logs (if available)."
            ],
            "01_by_cao": [
                "NOTES:",
                "This sheet summarizes truncated files by CAO number.",
                "",
                "cao_number: CAO number.",
                "n_truncated_files: Number of truncated files for this CAO.",
                "sectors: Sectors associated with this CAO (if available).",
                "sbi_codes: SBI codes associated with this CAO (if available).",
                "filenames: Sample filenames (first 3, or all if ≤3)."
            ],
            "02_by_sector": [
                "NOTES:",
                "This sheet summarizes truncated files by sector.",
                "",
                "sector: Sector name (or 'Unknown' if not available).",
                "n_truncated_files: Number of truncated files in this sector.",
                "n_unique_caos: Number of unique CAO numbers in this sector.",
                "cao_numbers: List of CAO numbers in this sector."
            ],
            "03_statistics": [
                "NOTES:",
                "This sheet provides overall statistics about truncated files (level 4 only).",
                "",
                "metric: Name of the statistic.",
                "value: Value of the statistic.",
                "",
                "Statistics include:",
                "  - Counts: Total files, unique CAOs",
                "  - File sizes: Mean, median, min, max for truncated files and original files",
                "  - Processing: Mean and median processing times",
                "  - Tokens: Mean and median input/output/total tokens"
            ],
            "04_file_status_comparison": [
                "NOTES:",
                "This sheet shows the status of all files:",
                "",
                "status categories:",
                "  - successfully_extracted: Files that appear in the salary CSV",
                "  - truncated_4_all_attempts_exhausted: Files in truncated_4 folder (all attempts failed)",
                "  - truncated_3_might_be_retried: Files in truncated_3 folder (might be retried with super compact schema)",
                "  - truncated_2_might_be_retried: Files in truncated_2 folder (might be retried with split extraction)",
                "  - truncated_1_might_be_retried: Files in truncated folder (might be retried with compact schema)",
                "  - missing_not_processed: Files in CAO info but not extracted and not in any truncated folder",
                "",
                "cao_number: CAO number",
                "filename: Original filename (normalized)",
                "truncation_level: Level of truncation (1-4) if applicable",
                "was_saved_after_retry: True if file was successfully saved after retry (exists in outputs/llm_analysis/salary), False otherwise",
                "",
                "IMPORTANT: If was_saved_after_retry is True for files in truncated_1/2/3, those files were",
                "successfully extracted with compact/split/super compact schemas and should be in the Excel output."
            ],
            "05_summary_counts": [
                "NOTES:",
                "This sheet provides summary counts by file status.",
                "",
                "status: File status category",
                "n_files: Number of files in this category",
                "n_unique_ids: Number of unique IDs in this category",
                "n_saved_after_retry: Number of files in this category that were successfully saved after retry",
                "",
                "This helps answer:",
                "  - How many files were successfully extracted?",
                "  - How many files failed at each truncation level?",
                "  - How many files in truncated_1/2/3 were successfully saved after retry?",
                "  - How many files are missing/never processed?",
                "  - What is the total expected from CAO info?",
                "",
                "IMPORTANT: If n_saved_after_retry > 0 for truncated_1/2/3, those files were",
                "successfully extracted with compact/split/super compact schemas and should be in the Excel output."
            ]
        }
        
        # Add notes to sheets
        for sheet_name, note_lines in notes.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                max_row = ws.max_row
                note_start_row = max_row + 3
                
                for i, note_line in enumerate(note_lines):
                    cell = ws.cell(row=note_start_row + i, column=1)
                    cell.value = note_line
                    try:
                        existing_font = cell.font
                        cell.font = Font(italic=True, name=existing_font.name if existing_font else 'Calibri',
                                       size=existing_font.size if existing_font else 11,
                                       bold=existing_font.bold if existing_font else False,
                                       color=existing_font.color if existing_font and existing_font.color else None)
                    except Exception:
                        cell.font = Font(italic=True)
        
        wb.save(OUTPUT_EXCEL)
        print(f"\n✓ Successfully created Excel workbook with {len(sheets)} sheets and explanatory notes")
        print(f"  Output: {OUTPUT_EXCEL}")
    except Exception as e:
        print(f"\n  ERROR: Could not write Excel file: {e}")
        import traceback
        traceback.print_exc()
        return


if __name__ == "__main__":
    main()

