#!/usr/bin/env python3
"""
Rebuild Summary sheet comparison output with formatting-tolerant equality.

PURPOSE:
    This one-off utility recalculates only the `fields_changed` column in the
    `Summary` worksheet of an amenities comparison workbook. It is designed for
    cases where comparison noise is caused by value formatting differences
    (for example `8` vs `8.0`, `true` vs `True`, or trailing whitespace).

WHAT THIS SCRIPT CHANGES:
    - Recomputes `Summary.fields_changed` using canonical value comparison.
    - Writes updates only to non-header `fields_changed` cells in `Summary`.

WHAT THIS SCRIPT DOES NOT CHANGE:
    - Does not modify `extracted_amenities`.
    - Does not modify `extracted_data_non_salary`.
    - Does not rewrite formulas in the Summary test/formula area.

DEFAULT INPUT/OUTPUT:
    - Input:
      `outputs/excel/ClaudeCodeTest/extracted_amenities.xlsx`
    - Output copy:
      `outputs/excel/ClaudeCodeTest/extracted_amenities.summary_fixed.xlsx`

USAGE:
    conda run -n caos-extract python scripts/validation/rebuild_summary_fields_changed.py

    # Optional custom paths:
    conda run -n caos-extract python scripts/validation/rebuild_summary_fields_changed.py \
      --input outputs/excel/ClaudeCodeTest/extracted_amenities.xlsx \
      --output outputs/excel/ClaudeCodeTest/extracted_amenities.summary_fixed.xlsx
"""

from __future__ import annotations

import argparse
import math
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
DEFAULT_INPUT = PROJECT_ROOT / "outputs/excel/ClaudeCodeTest/extracted_amenities.xlsx"
DEFAULT_OUTPUT = (
    PROJECT_ROOT / "outputs/excel/ClaudeCodeTest/extracted_amenities.summary_fixed.xlsx"
)

SUMMARY_SHEET = "Summary"
CLAUDE_SHEET = "extracted_amenities"
GEMINI_SHEET = "extracted_data_non_salary"
SUMMARY_HEADER_ROW = 2
SOURCE_HEADER_ROW = 1
SUMMARY_DATA_START_ROW = 3
SOURCE_DATA_START_ROW = 2
MAX_CONSECUTIVE_EMPTY_ROWS = 2000


@dataclass
class RebuildStats:
    """
    Hold summary metrics for rebuild operations.

    Attributes:
        rows_processed: Number of Summary data rows inspected.
        rows_updated: Number of rows where `fields_changed` was modified.
        fields_before_total: Total `fields_changed` entries before rebuild.
        fields_after_total: Total `fields_changed` entries after rebuild.
        formatting_only_removed: Count of original changed fields that became equal
            after canonical formatting-tolerant comparison.
        row_ids_missing_in_sources: Number of Summary IDs not found in one or both
            source sheets.
    """

    rows_processed: int = 0
    rows_updated: int = 0
    fields_before_total: int = 0
    fields_after_total: int = 0
    formatting_only_removed: int = 0
    row_ids_missing_in_sources: int = 0


def parse_args() -> argparse.Namespace:
    """
    Parse command-line arguments for input/output workbook paths.

    Returns:
        Parsed CLI namespace with `input` and `output` paths.
    """

    parser = argparse.ArgumentParser(
        description="Rebuild Summary.fields_changed with formatting-tolerant comparison."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help=f"Path to source workbook (default: {DEFAULT_INPUT})",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Path to output workbook copy (default: {DEFAULT_OUTPUT})",
    )
    parser.add_argument(
        "--max-data-rows",
        type=int,
        default=5000,
        help=(
            "Maximum number of data rows to scan per relevant sheet "
            "(default: 5000, suitable for expected ~1.5k rows)."
        ),
    )
    return parser.parse_args()


def build_header_index(header_values: Sequence[Any]) -> Dict[str, int]:
    """
    Create a mapping from header name to zero-based column index.

    Args:
        header_values: Row values containing column headers.

    Returns:
        Dictionary mapping string header names to their column index.
    """

    return {
        str(name): idx
        for idx, name in enumerate(header_values)
        if name is not None and str(name).strip() != ""
    }


def parse_changed_fields(raw_value: Any) -> List[str]:
    """
    Parse a pipe-delimited `fields_changed` cell into an ordered field list.

    Args:
        raw_value: Raw cell value from Summary `fields_changed`.

    Returns:
        Ordered list of non-empty field names.
    """

    if raw_value is None:
        return []
    value = str(raw_value).strip()
    if not value:
        return []
    return [item.strip() for item in value.split("|") if item.strip()]


def normalize_number(text_value: str) -> Decimal | None:
    """
    Parse a numeric string into Decimal for canonical numeric comparison.

    Args:
        text_value: Candidate numeric string.

    Returns:
        A Decimal when parsing succeeds, otherwise None.
    """

    try:
        return Decimal(text_value)
    except (InvalidOperation, ValueError):
        return None


def canonicalize_value(value: Any) -> Any:
    """
    Convert a raw cell value into a canonical representation for equality checks.

    Args:
        value: Original value from Excel.

    Returns:
        Canonical value where formatting-only differences collapse to one form.
        Preserves semantic differences.
    """

    if value is None:
        return None

    if isinstance(value, bool):
        return value

    if isinstance(value, float):
        if math.isnan(value):
            return None
        return Decimal(str(value)).normalize()

    if isinstance(value, int):
        return Decimal(value).normalize()

    if isinstance(value, Decimal):
        return value.normalize()

    if isinstance(value, str):
        stripped = value.strip()
        if stripped == "":
            return None

        lowered = stripped.lower()
        if lowered == "true":
            return True
        if lowered == "false":
            return False

        parsed_number = normalize_number(stripped)
        if parsed_number is not None:
            return parsed_number.normalize()

        return stripped

    return value


def worksheet_rows_by_id(
    worksheet: Any,
    header_index: Dict[str, int],
    start_row: int,
    max_data_rows: int,
    max_consecutive_empty: int = MAX_CONSECUTIVE_EMPTY_ROWS,
) -> Dict[Any, Sequence[Any]]:
    """
    Build a lookup map from row `id` to complete row values for a worksheet.

    Args:
        worksheet: Openpyxl worksheet containing row data.
        header_index: Header-to-index map for the worksheet.
        start_row: First data row index (1-based).
        max_data_rows: Hard cap for scanned data rows.
        max_consecutive_empty: Early-stop threshold for trailing empty rows.

    Returns:
        Dictionary mapping ID values to row tuples.
    """

    id_idx = header_index["id"]
    lookup: Dict[Any, Sequence[Any]] = {}
    consecutive_empty = 0
    max_row = start_row + max_data_rows - 1
    for row_values in worksheet.iter_rows(
        min_row=start_row, max_row=max_row, values_only=True
    ):
        row_id = row_values[id_idx]
        if row_id is None:
            consecutive_empty += 1
            if lookup and consecutive_empty >= max_consecutive_empty:
                break
            continue
        consecutive_empty = 0
        lookup[row_id] = row_values
    return lookup


def iter_common_fields(
    claude_header_idx: Dict[str, int], gemini_header_idx: Dict[str, int]
) -> Iterable[str]:
    """
    Yield source field names available in both extraction worksheets.

    Args:
        claude_header_idx: Header map for `extracted_amenities`.
        gemini_header_idx: Header map for `extracted_data_non_salary`.

    Yields:
        Shared field names sorted for deterministic output order.
    """

    excluded = {"id"}
    shared = set(claude_header_idx.keys()) & set(gemini_header_idx.keys())
    for field_name in sorted(shared):
        if field_name in excluded:
            continue
        yield field_name


def rebuild_summary_fields_changed(
    workbook_path: Path, output_path: Path, max_data_rows: int
) -> RebuildStats:
    """
    Recompute Summary `fields_changed` using canonicalized field comparison.

    Args:
        workbook_path: Source workbook path.
        output_path: Output workbook copy path to write.

    Returns:
        Rebuild statistics describing before/after differences.
    """

    workbook = load_workbook(workbook_path)
    ws_summary = workbook[SUMMARY_SHEET]
    ws_claude = workbook[CLAUDE_SHEET]
    ws_gemini = workbook[GEMINI_SHEET]

    summary_headers = [
        cell.value
        for cell in next(
            ws_summary.iter_rows(
                min_row=SUMMARY_HEADER_ROW,
                max_row=SUMMARY_HEADER_ROW,
                values_only=False,
            )
        )
    ]
    source_headers_claude = [
        cell.value
        for cell in next(
            ws_claude.iter_rows(
                min_row=SOURCE_HEADER_ROW, max_row=SOURCE_HEADER_ROW, values_only=False
            )
        )
    ]
    source_headers_gemini = [
        cell.value
        for cell in next(
            ws_gemini.iter_rows(
                min_row=SOURCE_HEADER_ROW, max_row=SOURCE_HEADER_ROW, values_only=False
            )
        )
    ]

    summary_idx = build_header_index(summary_headers)
    claude_idx = build_header_index(source_headers_claude)
    gemini_idx = build_header_index(source_headers_gemini)

    required_summary = {"id", "fields_changed"}
    missing_summary = sorted(required_summary - set(summary_idx.keys()))
    if missing_summary:
        raise KeyError(f"Summary missing required columns: {missing_summary}")
    if "id" not in claude_idx or "id" not in gemini_idx:
        raise KeyError("Source sheets must both contain `id` column.")

    claude_by_id = worksheet_rows_by_id(
        ws_claude, claude_idx, SOURCE_DATA_START_ROW, max_data_rows=max_data_rows
    )
    gemini_by_id = worksheet_rows_by_id(
        ws_gemini, gemini_idx, SOURCE_DATA_START_ROW, max_data_rows=max_data_rows
    )
    shared_fields = list(iter_common_fields(claude_idx, gemini_idx))

    stats = RebuildStats()
    id_col = summary_idx["id"] + 1
    changed_col = summary_idx["fields_changed"] + 1

    consecutive_empty_summary = 0
    summary_max_row = SUMMARY_DATA_START_ROW + max_data_rows - 1
    for row_number in range(SUMMARY_DATA_START_ROW, summary_max_row + 1):
        summary_id = ws_summary.cell(row=row_number, column=id_col).value
        if summary_id is None:
            consecutive_empty_summary += 1
            if stats.rows_processed > 0 and consecutive_empty_summary >= MAX_CONSECUTIVE_EMPTY_ROWS:
                break
            continue
        consecutive_empty_summary = 0

        stats.rows_processed += 1
        old_changed_raw = ws_summary.cell(row=row_number, column=changed_col).value
        old_fields = parse_changed_fields(old_changed_raw)
        stats.fields_before_total += len(old_fields)

        claude_row = claude_by_id.get(summary_id)
        gemini_row = gemini_by_id.get(summary_id)
        if claude_row is None or gemini_row is None:
            stats.row_ids_missing_in_sources += 1
            continue

        new_fields: List[str] = []
        for field_name in shared_fields:
            left = claude_row[claude_idx[field_name]]
            right = gemini_row[gemini_idx[field_name]]
            if canonicalize_value(left) != canonicalize_value(right):
                new_fields.append(field_name)

        new_changed_value = "|".join(new_fields) if new_fields else None
        stats.fields_after_total += len(new_fields)

        old_fields_set = set(old_fields)
        new_fields_set = set(new_fields)
        removed_fields = old_fields_set - new_fields_set
        for field_name in removed_fields:
            left = claude_row[claude_idx.get(field_name)] if field_name in claude_idx else None
            right = gemini_row[gemini_idx.get(field_name)] if field_name in gemini_idx else None
            if canonicalize_value(left) == canonicalize_value(right):
                stats.formatting_only_removed += 1

        if old_changed_raw != new_changed_value:
            ws_summary.cell(row=row_number, column=changed_col).value = new_changed_value
            stats.rows_updated += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return stats


def main() -> None:
    """
    Execute one-off Summary `fields_changed` rebuild and print metrics.

    Returns:
        None.
    """

    args = parse_args()
    input_path = args.input.resolve()
    output_path = args.output.resolve()

    if not input_path.is_file():
        raise FileNotFoundError(f"Input workbook not found: {input_path}")

    stats = rebuild_summary_fields_changed(
        input_path, output_path, max_data_rows=args.max_data_rows
    )

    print(f"Input workbook:  {input_path}")
    print(f"Output workbook: {output_path}")
    print(f"Rows processed: {stats.rows_processed}")
    print(f"Rows updated: {stats.rows_updated}")
    print(f"fields_changed total before: {stats.fields_before_total}")
    print(f"fields_changed total after:  {stats.fields_after_total}")
    print(f"Formatting-only changed fields removed: {stats.formatting_only_removed}")
    print(f"Summary IDs missing in source sheets: {stats.row_ids_missing_in_sources}")


if __name__ == "__main__":
    main()
