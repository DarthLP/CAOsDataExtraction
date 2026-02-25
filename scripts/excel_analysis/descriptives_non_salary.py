"""
CAO Non-Salary Descriptives Script

This script reads the non-salary CSV output and produces a multi-sheet Excel workbook
with comprehensive descriptive statistics including variable health, sample overview,
domain coverage, headline features, numeric intensities, and modern policy comparisons.

USAGE:
    python scripts/excel_analysis/descriptives_non_salary.py

INPUT:
    - outputs/excel/new_results/extracted_data_non_salary.csv

OUTPUT:
    - outputs/analysis/non_salary_descriptives.xlsx (multi-sheet Excel workbook)
"""

import os
import sys
from pathlib import Path
from typing import Dict, List, Any, Optional, Union
import pandas as pd
import numpy as np
from datetime import datetime

# Add the parent directory to Python path so we can import utils
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

# Import analysis utilities
from scripts.excel_analysis.analysis_utils import (
    calculate_descriptive_stats, create_boolean_summary, parse_cao_date_series
)

# =============================================================================
# PATH CONFIGURATION
# =============================================================================
INPUT_CSV_PATH = "outputs/excel/new_results/extracted_data_non_salary.csv"
OUTPUT_EXCEL_PATH = "outputs/analysis/non_salary_descriptives.xlsx"

# =============================================================================
# DOMAIN & FEATURE DEFINITIONS
# =============================================================================

DOMAIN_FLAGS = {
    "bonus": "bonus_has_bonus_schemes",
    "pension": "pension_has_pension_scheme",
    "leave": "leave_has_leave_enhancements",
    "termination": "term_has_termination_rules",
    "overtime": "overtime_has_overtime_rules",
    "training": "training_has_training_rights",
    "homeoffice": "homeoffice_has_homeoffice_rights",
    "contract": "contract_has_contract_type_rules",
    "fringe": "fringe_has_fringe_benefits",
    "safety": [
        "safety_harassment_protocol_present",
        "safety_integrity_protocol_present",
        "safety_confidential_counsellor_present",
        "safety_safety_training_present",
        "safety_arbodienst_access_provided",
        "safety_wellbeing_program_present",
    ],
    "childcare": [
        "childcare_childcare_support_present",
        "childcare_inhouse_present",
        "childcare_discount_present",
        "childcare_priority_access",
    ],
    "ai": "ai_ai_policy_exists",
}

HEADLINE_FEATURES = {
    "bonus": [
        "bonus_has_bonus_schemes",
        "bonus_sign_on_bonus_present",
        "bonus_thirteenth_month",
        "bonus_profit_sharing_present",
        "bonus_performance_bonus_present",
        "bonus_job_allowances_present",
        "bonus_qual_bonus_present",
        "bonus_seniority_loyalty_bonus",
        "bonus_retire_gratuity_present",
    ],
    "wage_scales": [
        "wage_entry_step_exp_present",
        "wage_pers_allow_max_scale",
        "wage_perf_step_var_present",
    ],
    "pension": [
        "pension_has_pension_scheme",
        "pension_mandatory_participation",
        "pension_accrual_stat_leaves",
        "pension_accrual_illness_y2",
        "pension_excedent_present",
        "pension_premium_eq_split",
        "pension_hetero_pension",
    ],
    "leave": [
        "leave_has_leave_enhancements",
        "leave_has_above_statutory_maternity",
        "leave_paternity_explicitly_above_statutory",
        "leave_parental_statutory_ref",
        "leave_parental_exceptions",
        "leave_parental_eligibility_present",
        "leave_parental_topup_present",
        "leave_abortion_present",
        "leave_sick_topup_present",
        "leave_sickpay_extra_insurance_present",
        "leave_care_statutory_ref",
        "leave_care_exceptions",
        "leave_care_topup_present",
        "leave_hetero_present",
        "leave_liberation_day_annual",
        "leave_liberation_day_lustrum",
        "leave_extra_seniority_present",
    ],
    "termination": [
        "term_has_termination_rules",
        "term_hetero_present",
        "term_notice_tenure_present",
        "term_shorten_notice_uwv",
        "term_sick_dismissal_prot",
        "term_end_at_AOW_auto",
        "term_probation_allowed",
        "term_severance_ww_supplement",
    ],
    "overtime": [
        "overtime_has_overtime_rules",
        "overtime_hetero_present",
        "overtime_shift_allowance_present",
    ],
    "training": [
        "training_has_training_rights",
        "training_fund_present",
        "training_reclaim_clause_present",
        "training_mandatory_training_paid",
    ],
    "homeoffice": [
        "homeoffice_has_homeoffice_rights",
        "homeoffice_stipend_present",
        "homeoffice_costs_reimbursed",
        "homeoffice_agreement_required",
        "homeoffice_health_safety_guarantee",
        "homeoffice_travel_time_compensation",
    ],
    "contract": [
        "contract_has_contract_type_rules",
        "contract_part_time_allowed",
        "contract_minmax_hours_contract_allowed",
        "contract_zero_hour_oncall_allowed",
        "contract_ketenregeling_deviation_present",
        "contract_conversion_rights_temp_to_perm_present",
        "contract_workhours_adjustment_right_present",
    ],
    "fringe": [
        "fringe_has_fringe_benefits",
        "fringe_commuting_allowance_present",
        "fringe_bike_scheme_present",
        "fringe_internet_or_phone_reimbursement_present",
        "fringe_meal_benefit_present",
        "fringe_health_insurance_support_present",
        "fringe_insurance_or_savings_benefit_present",
        "fringe_relocation_allowance_present",
        "fringe_mandatory_certifications_paid",
    ],
    "safety": [
        "safety_harassment_protocol_present",
        "safety_integrity_protocol_present",
        "safety_confidential_counsellor_present",
        "safety_reporting_channel_external",
        "safety_safety_training_present",
        "safety_safety_committee_present",
        "safety_rie_psa_required",
        "safety_psa_prevention_measures_present",
        "safety_arbodienst_access_provided",
        "safety_preventive_medical_checkup_present",
        "safety_workload_monitoring_present",
        "safety_wellbeing_program_present",
    ],
    "childcare": [
        "childcare_childcare_support_present",
        "childcare_inhouse_present",
        "childcare_discount_present",
        "childcare_priority_access",
        "childcare_funding_sector_fund",
    ],
    "ai": [
        "ai_ai_policy_exists",
        "ai_ai_governance_body_present",
        "ai_ai_training_rights_present",
    ],
}

HEADLINE_CATEGORICAL = {
    "pension": [
        "pension_pension_type",
        "pension_selection_rule_pension",
    ],
    "termination": [
        "term_dismissal_approval",
    ],
    "overtime": [
        "overtime_compensation_mode",
        "overtime_stacking_rule",
    ],
    "homeoffice": [
        "homeoffice_discretion",
    ],
    "fringe": [
        "fringe_meal_benefit_type",
    ],
    "childcare": [
        "childcare_provider_scope",
        "childcare_public_coord",
    ],
    "ai": [
        "ai_ai_automated_decisions",
    ],
}

NUMERIC_VARS = [
    "contract_full_time_hours_value",
    "leave_vacation_time_value",
    "leave_vacation_bonus_value",
    "leave_sickpay_duration_value",
    "leave_sickpay_continuation_value",
    "leave_paid_maternity_value",
    "leave_paid_paternity_value",
    "leave_parental_topup_pay_value",
    "leave_long_term_care_value",
    "leave_long_term_care_pay_value",
    "pension_employee_contrib_value",
    "pension_accrual_rate_value",
    "pension_franchise_value",
    "pension_retire_age_normal_value",
    "term_employee_notice_value",
    "term_employer_notice_value",
    "term_probation_fixedterm_value",
    "term_probation_indef_value",
    "overtime_allowance_value",
    "overtime_shift_allowance_range_min",
    "overtime_shift_allowance_range_max",
    "overtime_min_rest_between_shifts_value",
    "overtime_max_hours_per_day_value",
    "overtime_max_hours_per_week_value",
    "overtime_compulsory_annual_value",
    "training_time_yearly_value",
    "training_budget_value",
    "training_cost_reimbursement_value",
    "fringe_commuting_allowance_value",
    "fringe_meal_benefit_amt_value",
    "fringe_relocation_allowance_value",
    "childcare_childcare_support_value",
    "childcare_childcare_support_cap_value",
    "childcare_age_min_value",
    "childcare_age_max_value",
    "homeoffice_entitlement_value",
    "homeoffice_stipend_value",
]

MODERN_FEATURES = [
    "homeoffice_has_homeoffice_rights",
    "homeoffice_stipend_present",
    "ai_ai_policy_exists",
    "childcare_childcare_support_present",
    "training_fund_present",
    "leave_sick_topup_present",
    "leave_parental_topup_present",
    "safety_harassment_protocol_present",
    "safety_wellbeing_program_present",
]

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def normalize_boolean(series: pd.Series) -> pd.Series:
    """
    Normalize boolean values to True/False.
    
    Args:
        series: Series with potentially mixed boolean representations
        
    Returns:
        Series with normalized boolean values
    """
    result = series.copy()
    
    # Map common boolean representations
    bool_map = {
        1: True, 0: False,
        "1": True, "0": False,
        "yes": True, "no": False,
        "Yes": True, "No": False,
        "YES": True, "NO": False,
        "true": True, "false": False,
        "True": True, "False": False,
        "TRUE": True, "FALSE": False,
    }
    
    # Apply mapping where applicable
    for val, bool_val in bool_map.items():
        result = result.replace(val, bool_val)
    
    # Convert to boolean, keeping NaN
    result = result.astype(object)
    result = result.where(result.isin([True, False, np.nan]), np.nan)
    
    return result


def infer_var_type(series: pd.Series, name: str) -> str:
    """
    Infer variable type from series and column name.
    
    Args:
        series: Pandas Series
        name: Column name
        
    Returns:
        Type string: "boolean", "numeric", "date", "categorical", "text", "id/other"
    """
    # Check for boolean
    if series.dtype == bool:
        return "boolean"
    
    # Check if values collapse to boolean
    non_null = series.dropna()
    if len(non_null) > 0:
        unique_vals = set(non_null.unique())
        bool_like = {True, False, 1, 0, "1", "0", "yes", "no", "Yes", "No", 
                     "YES", "NO", "true", "false", "True", "False", "TRUE", "FALSE"}
        if unique_vals.issubset(bool_like):
            return "boolean"
    
    # Check for date
    if "date" in name.lower() or "datum" in name.lower():
        return "date"
    if pd.api.types.is_datetime64_any_dtype(series):
        return "date"
    
    # Check for text patterns in column name
    text_patterns = ["_note", "_rule", "_schedule", "_description", "_terms", "_scope_note", 
                     "_exclusions_note", "_backpay_terms", "_interest_or_surcharge"]
    if any(name.endswith(pattern) for pattern in text_patterns):
        return "text"
    
    # Try numeric
    numeric_series = pd.to_numeric(series, errors='coerce')
    non_null_numeric = numeric_series.dropna()
    if len(non_null_numeric) > 0:
        # If ≥20% of non-null values are numeric, consider it numeric
        if len(non_null_numeric) >= 0.2 * len(non_null):
            return "numeric"
    
    # Distinguish categorical from text
    n_distinct = non_null.nunique()
    
    # If many distinct values, likely text or id/other
    if n_distinct > 50:
        return "text"
    
    # For fields with moderate distinct count, analyze characteristics
    if n_distinct > 0:
        # Convert to string for analysis
        str_values = non_null.astype(str)
        
        # Calculate average string length
        avg_length = str_values.str.len().mean()
        
        # Check for text-like characteristics
        has_spaces = str_values.str.contains(' ', regex=False).sum() > 0
        has_punctuation = str_values.str.contains(r'[.,;:!?]', regex=True).sum() > 0
        has_long_values = (str_values.str.len() > 50).sum() > 0
        
        # Categorical fields typically:
        # - Short values (avg length < 20)
        # - No spaces or punctuation (or very few)
        # - Consistent length
        # - Values look like categories (underscores, short codes, etc.)
        
        # Text fields typically:
        # - Longer values (avg length > 20)
        # - Contain spaces and punctuation
        # - Variable length
        
        # If average length is long or has text characteristics, it's text
        if avg_length > 30 or (avg_length > 15 and (has_spaces or has_punctuation)) or has_long_values:
            return "text"
        
        # If distinct count is very low (<= 10) and values are short, likely categorical
        if n_distinct <= 10 and avg_length < 20:
            return "categorical"
        
        # If distinct count is moderate (11-50) and values are short, check more carefully
        if n_distinct <= 50:
            # Check if values look like categories (short, no spaces, consistent)
            length_std = str_values.str.len().std()
            # If values are short, consistent length, and few have spaces, likely categorical
            if avg_length < 15 and length_std < 5 and not has_spaces:
                return "categorical"
            else:
                return "text"
    
    # Default
    return "id/other"


def describe_numeric(series: pd.Series) -> Dict[str, Any]:
    """
    Describe numeric series with key statistics.
    
    Args:
        series: Pandas Series
        
    Returns:
        Dictionary with n, mean, std, min, p25, median, p75, max
    """
    numeric_series = pd.to_numeric(series, errors='coerce')
    non_null = numeric_series.dropna()
    
    if len(non_null) == 0:
        return {
            'n': 0,
            'mean': np.nan,
            'std': np.nan,
            'min': np.nan,
            'p25': np.nan,
            'median': np.nan,
            'p75': np.nan,
            'max': np.nan
        }
    
    return {
        'n': len(non_null),
        'mean': non_null.mean(),
        'std': non_null.std(),
        'min': non_null.min(),
        'p25': non_null.quantile(0.25),
        'median': non_null.median(),
        'p75': non_null.quantile(0.75),
        'max': non_null.max()
    }


def build_latest_cao_view(df: pd.DataFrame, cao_col: str = "cao_number", 
                         date_col: str = "ingangsdatum") -> pd.DataFrame:
    """
    Build latest CAO view - one row per cao_number with the most recent date.
    
    Args:
        df: Full DataFrame
        cao_col: Column name for CAO number
        date_col: Column name for date
        
    Returns:
        DataFrame with one row per cao_number (latest date)
    """
    if len(df) == 0:
        print(f"  Warning: Input DataFrame is empty")
        return pd.DataFrame()
    
    if cao_col not in df.columns:
        print(f"  Warning: Column '{cao_col}' not found. Available columns: {list(df.columns[:10])}...")
        # Try alternative column names
        alt_names = ['CAO_NUMBER', 'cao', 'CAO', 'cao_num']
        for alt in alt_names:
            if alt in df.columns:
                print(f"  Using alternative column '{alt}' instead")
                cao_col = alt
                break
        else:
            print(f"  Returning empty DataFrame.")
            return pd.DataFrame()
    
    if date_col not in df.columns:
        print(f"  Warning: Column '{date_col}' not found. Available columns: {list(df.columns[:10])}...")
        return pd.DataFrame()
    
    # Parse date column (CAO metadata dates are in DD/MM/YYYY format)
    df_copy = df.copy()
    dayfirst = date_col in ['ingangsdatum', 'expiratiedatum', 'datum_kennisgeving']
    df_copy[date_col] = parse_cao_date_series(df_copy[date_col], dayfirst=dayfirst)
    
    # Check if we have valid dates
    valid_dates = df_copy[date_col].notna()
    if valid_dates.sum() == 0:
        print(f"  Warning: No valid dates found in '{date_col}' after parsing")
        return pd.DataFrame()
    
    # Check if we have valid CAO numbers
    valid_caos = df_copy[cao_col].notna()
    if valid_caos.sum() == 0:
        print(f"  Warning: No valid CAO numbers found in '{cao_col}'")
        return pd.DataFrame()
    
    # Filter to rows with both valid CAO and date
    valid_mask = valid_caos & valid_dates
    if valid_mask.sum() == 0:
        print(f"  Warning: No rows with both valid CAO number and date")
        return pd.DataFrame()
    
    df_copy = df_copy[valid_mask].copy()
    
    # Reset index to create a column we can sort by for tie-breaking
    df_copy = df_copy.reset_index(drop=True)
    df_copy['_sort_index'] = df_copy.index
    
    # Sort by cao_number, date, and original index to break ties deterministically
    df_copy = df_copy.sort_values([cao_col, date_col, '_sort_index'], 
                                  na_position='last')
    
    # Keep last row per cao_number (which has max date due to sorting)
    try:
        df_latest = df_copy.groupby(cao_col, as_index=False).last()
        # Drop the temporary sort index column
        df_latest = df_latest.drop(columns=['_sort_index'], errors='ignore')
        print(f"  Successfully built latest view: {len(df_latest)} unique CAOs from {len(df_copy)} rows")
    except Exception as e:
        print(f"  Error in groupby: {e}")
        return pd.DataFrame()
    
    return df_latest


def compute_domain_present(df: pd.DataFrame, domain_flag: Union[str, List[str]]) -> pd.Series:
    """
    Compute domain presence indicator.
    
    Args:
        df: DataFrame
        domain_flag: Either a single column name (str) or list of column names
        
    Returns:
        Series with True/False/NaN indicating domain presence
    """
    if isinstance(domain_flag, str):
        # Single column
        if domain_flag not in df.columns:
            return pd.Series([np.nan] * len(df), index=df.index)
        col = df[domain_flag]
        return normalize_boolean(col)
    else:
        # List of columns - OR logic
        if not domain_flag:
            return pd.Series([np.nan] * len(df), index=df.index)
        
        # Check which columns exist
        existing_cols = [col for col in domain_flag if col in df.columns]
        if not existing_cols:
            return pd.Series([np.nan] * len(df), index=df.index)
        
        # OR across existing columns
        result = pd.Series([False] * len(df), index=df.index)
        for col in existing_cols:
            bool_col = normalize_boolean(df[col])
            result = result | (bool_col == True)
        
        return result


# =============================================================================
# SHEET GENERATION FUNCTIONS
# =============================================================================

def create_variable_health_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create variable health sheet with one row per column.
    
    Args:
        df: Input DataFrame
        
    Returns:
        DataFrame with variable health statistics
    """
    total_rows = len(df)
    results = []
    
    for col_name in df.columns:
        series = df[col_name]
        non_null = series.dropna()
        n_nonmissing = len(non_null)
        share_nonmissing = n_nonmissing / total_rows if total_rows > 0 else 0
        n_distinct = non_null.nunique() if len(non_null) > 0 else 0
        
        var_type = infer_var_type(series, col_name)
        
        row = {
            'variable_name': col_name,
            'inferred_type': var_type,
            'n_nonmissing': n_nonmissing,
            'share_nonmissing': share_nonmissing,
            'n_distinct': n_distinct
        }
        
        # Type-specific statistics
        if var_type == "boolean":
            bool_series = normalize_boolean(series)
            n_true = (bool_series == True).sum()
            n_false = (bool_series == False).sum()
            n_missing = total_rows - n_nonmissing
            
            row.update({
                'n_true': n_true,
                'share_true': n_true / total_rows if total_rows > 0 else 0,
                'n_false': n_false,
                'share_false': n_false / total_rows if total_rows > 0 else 0,
                'n_missing': n_missing
            })
        
        elif var_type == "numeric":
            numeric_series = pd.to_numeric(series, errors='coerce')
            non_null_num = numeric_series.dropna()
            
            if len(non_null_num) > 0:
                row.update({
                    'mean': non_null_num.mean(),
                    'median': non_null_num.median(),
                    'min': non_null_num.min(),
                    'max': non_null_num.max(),
                    'p25': non_null_num.quantile(0.25),
                    'p75': non_null_num.quantile(0.75),
                    'share_zero': (non_null_num == 0).sum() / len(non_null_num) if len(non_null_num) > 0 else 0
                })
            else:
                row.update({
                    'mean': np.nan, 'median': np.nan, 'min': np.nan, 'max': np.nan,
                    'p25': np.nan, 'p75': np.nan, 'share_zero': np.nan
                })
        
        elif var_type == "categorical":
            if len(non_null) > 0:
                value_counts = non_null.value_counts()
                top_category = value_counts.index[0] if len(value_counts) > 0 else np.nan
                top_category_count = value_counts.iloc[0] if len(value_counts) > 0 else 0
                row.update({
                    'top_category': top_category,
                    'top_category_share': top_category_count / total_rows if total_rows > 0 else 0
                })
            else:
                row.update({
                    'top_category': np.nan,
                    'top_category_share': np.nan
                })
        
        elif var_type == "date":
            dayfirst = series.name in ['ingangsdatum', 'expiratiedatum', 'datum_kennisgeving']
            date_series = parse_cao_date_series(series, dayfirst=dayfirst)
            non_null_dates = date_series.dropna()
            if len(non_null_dates) > 0:
                row.update({
                    'min_date': non_null_dates.min(),
                    'max_date': non_null_dates.max()
                })
            else:
                row.update({
                    'min_date': np.nan,
                    'max_date': np.nan
                })
        
        elif var_type == "text":
            if len(non_null) > 0:
                char_lengths = non_null.astype(str).str.len()
                row.update({
                    'avg_char_length': char_lengths.mean()
                })
            else:
                row.update({
                    'avg_char_length': np.nan
                })
        
        results.append(row)
    
    return pd.DataFrame(results)


def create_unused_fields_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create sheet listing fields that are completely unused.
    
    For non-boolean fields: n_nonmissing = 0
    For boolean fields: n_true = 0 (all False or missing)
    
    Args:
        df: Input DataFrame
        
    Returns:
        DataFrame with unused fields and their types
    """
    total_rows = len(df)
    results = []
    
    for col_name in df.columns:
        series = df[col_name]
        non_null = series.dropna()
        n_nonmissing = len(non_null)
        
        var_type = infer_var_type(series, col_name)
        
        # Check if field is unused
        is_unused = False
        if var_type == "boolean":
            bool_series = normalize_boolean(series)
            n_true = (bool_series == True).sum()
            if n_true == 0:
                is_unused = True
        else:
            # For non-boolean fields, check if n_nonmissing = 0
            if n_nonmissing == 0:
                is_unused = True
        
        if is_unused:
            results.append({
                'variable_name': col_name,
                'inferred_type': var_type,
                'n_nonmissing': n_nonmissing,
                'n_true': (normalize_boolean(series) == True).sum() if var_type == "boolean" else np.nan,
                'n_false': (normalize_boolean(series) == False).sum() if var_type == "boolean" else np.nan,
                'n_missing': total_rows - n_nonmissing
            })
    
    return pd.DataFrame(results)


def create_sample_overview_sheet(df: pd.DataFrame, df_latest: pd.DataFrame) -> pd.DataFrame:
    """
    Create sample overview sheet with panel, cross-section, and retroactivity summaries.
    
    Args:
        df: Full DataFrame
        df_latest: Latest CAO view DataFrame
        
    Returns:
        DataFrame with sample overview statistics
    """
    results = []
    
    # a) Panel by start year
    if "ingangsdatum" in df.columns:
        df_copy = df.copy()
        df_copy["ingangsdatum"] = parse_cao_date_series(df_copy["ingangsdatum"], dayfirst=True)
        df_copy["start_year"] = df_copy["ingangsdatum"].dt.year
        
        for year, year_group in df_copy.groupby("start_year"):
            if pd.isna(year):
                continue
            
            year_row = {
                'section': 'panel_by_year',
                'year': int(year),
                'n_contracts': len(year_group),
            }
            
            if "cao_number" in year_group.columns:
                year_row['n_cao'] = year_group['cao_number'].nunique()
            else:
                year_row['n_cao'] = np.nan
            
            if "general_avv_applies" in year_group.columns:
                avv_bool = normalize_boolean(year_group["general_avv_applies"])
                year_row['share_avv'] = (avv_bool == True).mean()
            else:
                year_row['share_avv'] = np.nan
            
            results.append(year_row)
    
    # b) Cross-section (latest CAO)
    if len(df_latest) > 0:
        cross_section_row = {
            'section': 'cross_section_latest',
            'n_cao': len(df_latest)
        }
        
        if "general_cao_scope_type" in df_latest.columns:
            scope_counts = df_latest["general_cao_scope_type"].value_counts()
            for scope_type, count in scope_counts.items():
                cross_section_row[f'share_{scope_type}'] = count / len(df_latest)
                cross_section_row[f'n_{scope_type}'] = count
        
        if "general_sbi_code_primary" in df_latest.columns:
            sbi_codes = df_latest["general_sbi_code_primary"].dropna()
            cross_section_row['n_sbi_codes'] = sbi_codes.nunique()
        else:
            cross_section_row['n_sbi_codes'] = np.nan
        
        results.append(cross_section_row)
    
    # b2) CAO scope type distribution (separate section for clarity)
    if "general_cao_scope_type" in df.columns:
        scope_counts = df["general_cao_scope_type"].value_counts()
        for scope_type, count in scope_counts.items():
            scope_row = {
                'section': 'scope_type_distribution',
                'scope_type': scope_type,
                'n_cao': count,
                'share': count / len(df) if len(df) > 0 else 0
            }
            results.append(scope_row)
    
    # c) Retroactivity summary
    if "general_retro_applies" in df.columns:
        retro_bool = normalize_boolean(df["general_retro_applies"])
        share_retroactive = (retro_bool == True).mean()
        
        retro_row = {
            'section': 'retroactivity',
            'share_retroactive': share_retroactive
        }
        
        # Conditional on retroactive == True
        retro_true_mask = retro_bool == True
        retro_true_df = df[retro_true_mask]
        
        if len(retro_true_df) > 0:
            if "general_retro_backpay_due" in retro_true_df.columns:
                backpay_bool = normalize_boolean(retro_true_df["general_retro_backpay_due"])
                retro_row['share_backpay_due'] = (backpay_bool == True).mean()
            else:
                retro_row['share_backpay_due'] = np.nan
            
            # Retro length
            if ("general_retro_start_date" in retro_true_df.columns and 
                "general_retro_end_date" in retro_true_df.columns):
                start_dates = pd.to_datetime(retro_true_df["general_retro_start_date"], errors='coerce')
                end_dates = pd.to_datetime(retro_true_df["general_retro_end_date"], errors='coerce')
                retro_lengths = (end_dates - start_dates).dt.days
                retro_lengths_clean = retro_lengths.dropna()
                
                if len(retro_lengths_clean) > 0:
                    retro_row['retro_length_median'] = retro_lengths_clean.median()
                    retro_row['retro_length_p25'] = retro_lengths_clean.quantile(0.25)
                    retro_row['retro_length_p75'] = retro_lengths_clean.quantile(0.75)
                else:
                    retro_row['retro_length_median'] = np.nan
                    retro_row['retro_length_p25'] = np.nan
                    retro_row['retro_length_p75'] = np.nan
            else:
                retro_row['retro_length_median'] = np.nan
                retro_row['retro_length_p25'] = np.nan
                retro_row['retro_length_p75'] = np.nan
            
            if "general_retro_int_surcharge" in retro_true_df.columns:
                interest_col = retro_true_df["general_retro_int_surcharge"]
                retro_row['share_interest_or_surcharge'] = (interest_col.notna() & (interest_col != "")).mean()
            else:
                retro_row['share_interest_or_surcharge'] = np.nan
        else:
            retro_row['share_backpay_due'] = np.nan
            retro_row['retro_length_median'] = np.nan
            retro_row['retro_length_p25'] = np.nan
            retro_row['retro_length_p75'] = np.nan
            retro_row['share_interest_or_surcharge'] = np.nan
        
        results.append(retro_row)
    
    # d) Date comparison: ingangsdatum (website) vs general_start_date (PDF)
    if "ingangsdatum" in df.columns and "general_start_date" in df.columns:
        df_dates = df.copy()
        df_dates["ingangsdatum"] = parse_cao_date_series(df_dates["ingangsdatum"], dayfirst=True)
        df_dates["general_start_date"] = pd.to_datetime(
            df_dates["general_start_date"], errors='coerce'
        )
        
        # Filter to rows with both dates
        both_dates_mask = df_dates["ingangsdatum"].notna() & df_dates["general_start_date"].notna()
        df_both_dates = df_dates[both_dates_mask]
        
        if len(df_both_dates) > 0:
            # Calculate date differences (PDF date - website date)
            date_diff = (df_both_dates["general_start_date"] - df_both_dates["ingangsdatum"]).dt.days
            
            date_comp_row = {
                'section': 'date_comparison',
                'n_with_both_dates': len(df_both_dates),
                'n_exact_match': (date_diff == 0).sum(),
                'share_exact_match': (date_diff == 0).sum() / len(df_both_dates) if len(df_both_dates) > 0 else 0,
                'n_pdf_later_than_website': (date_diff > 0).sum(),
                'share_pdf_later_than_website': (date_diff > 0).sum() / len(df_both_dates) if len(df_both_dates) > 0 else 0,
                'n_pdf_earlier_than_website': (date_diff < 0).sum(),
                'share_pdf_earlier_than_website': (date_diff < 0).sum() / len(df_both_dates) if len(df_both_dates) > 0 else 0,
                'date_diff_days_median': date_diff.median(),
                'date_diff_days_mean': date_diff.mean(),
                'date_diff_days_p25': date_diff.quantile(0.25),
                'date_diff_days_p75': date_diff.quantile(0.75),
                'date_diff_days_min': date_diff.min(),
                'date_diff_days_max': date_diff.max()
            }
            
            # Count large discrepancies (>30 days)
            large_diff = (date_diff.abs() > 30)
            date_comp_row['n_large_discrepancy_gt_30_days'] = large_diff.sum()
            date_comp_row['share_large_discrepancy_gt_30_days'] = large_diff.sum() / len(df_both_dates) if len(df_both_dates) > 0 else 0
            
            results.append(date_comp_row)
    
    return pd.DataFrame(results)


def create_domain_coverage_latest_sheet(df_latest: pd.DataFrame) -> pd.DataFrame:
    """
    Create domain coverage sheet for latest CAO view.
    
    Shows for each domain how many CAOs (in the latest view) have that domain present.
    
    Args:
        df_latest: Latest CAO view DataFrame
        
    Returns:
        DataFrame with domain coverage statistics
    """
    if len(df_latest) == 0:
        print("  Warning: df_latest is empty, returning empty domain coverage sheet")
        return pd.DataFrame(columns=['domain', 'n_cao_with_domain', 'share_cao_with_domain'])
    
    results = []
    n_total = len(df_latest)
    
    for domain, domain_flag in DOMAIN_FLAGS.items():
        domain_present = compute_domain_present(df_latest, domain_flag)
        
        # Count only non-NaN values where domain_present == True
        # Skip NaN values in the calculation
        valid_mask = domain_present.notna()
        if valid_mask.sum() > 0:
            n_with_domain = (domain_present[valid_mask] == True).sum()
            # Share is calculated over total CAOs (not just valid ones)
            share_with_domain = n_with_domain / n_total if n_total > 0 else 0
        else:
            # All values are NaN (columns don't exist)
            n_with_domain = 0
            share_with_domain = 0.0
        
        results.append({
            'domain': domain,
            'n_cao_with_domain': n_with_domain,
            'share_cao_with_domain': share_with_domain
        })
    
    result_df = pd.DataFrame(results)
    
    # Check if all values are 0 (might indicate missing columns)
    if len(result_df) > 0 and (result_df['n_cao_with_domain'] == 0).all():
        print(f"  Warning: All domain coverage values are 0. This might indicate missing domain columns in the data.")
        print(f"  Expected columns like: bonus_has_bonus_schemes, pension_has_pension_scheme, etc.")
    
    return result_df


def create_domain_coverage_by_year_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create domain coverage by year sheet.
    
    Args:
        df: Full DataFrame
        
    Returns:
        DataFrame with domain coverage by year
    """
    if "ingangsdatum" not in df.columns:
        return pd.DataFrame(columns=['start_year', 'domain', 'share_with_domain', 'n_rows_year'])
    
    df_copy = df.copy()
    df_copy["ingangsdatum"] = parse_cao_date_series(df_copy["ingangsdatum"], dayfirst=True)
    df_copy["start_year"] = df_copy["ingangsdatum"].dt.year
    
    results = []
    
    for year, year_group in df_copy.groupby("start_year"):
        if pd.isna(year):
            continue
        
        year = int(year)
        n_rows_year = len(year_group)
        
        for domain, domain_flag in DOMAIN_FLAGS.items():
            domain_present = compute_domain_present(year_group, domain_flag)
            # Skip NaNs in mean calculation
            valid_mask = domain_present.notna()
            if valid_mask.sum() > 0:
                share_with_domain = (domain_present[valid_mask] == True).mean()
            else:
                share_with_domain = np.nan
            
            results.append({
                'start_year': year,
                'domain': domain,
                'share_with_domain': share_with_domain,
                'n_rows_year': n_rows_year
            })
    
    return pd.DataFrame(results)


def create_headline_features_latest_sheet(df_latest: pd.DataFrame) -> pd.DataFrame:
    """
    Create headline features sheet for latest CAO view.
    
    Shows headline policy features (boolean and categorical) for the latest CAO view.
    
    Args:
        df_latest: Latest CAO view DataFrame
        
    Returns:
        DataFrame with headline features statistics
    """
    if len(df_latest) == 0:
        print("  Warning: df_latest is empty, returning empty headline features sheet")
        return pd.DataFrame()
    
    results = []
    n_total = len(df_latest)
    
    # Boolean features
    for domain, features in HEADLINE_FEATURES.items():
        for feature in features:
            if feature not in df_latest.columns:
                continue
            
            bool_series = normalize_boolean(df_latest[feature])
            n_true = (bool_series == True).sum()
            n_false = (bool_series == False).sum()
            n_missing = bool_series.isna().sum()
            
            results.append({
                'domain': domain,
                'feature': feature,
                'type': 'boolean',
                'share_true': n_true / n_total if n_total > 0 else 0,
                'n_true': n_true,
                'n_false': n_false,
                'n_missing': n_missing
            })
    
    # Categorical features
    for domain, features in HEADLINE_CATEGORICAL.items():
        for feature in features:
            if feature not in df_latest.columns:
                continue
            
            series = df_latest[feature]
            non_null = series.dropna()
            n_nonmissing = len(non_null)
            n_distinct = non_null.nunique() if len(non_null) > 0 else 0
            
            if len(non_null) > 0:
                value_counts = non_null.value_counts()
                top_category = value_counts.index[0] if len(value_counts) > 0 else np.nan
                top_category_count = value_counts.iloc[0] if len(value_counts) > 0 else 0
                top_category_share = top_category_count / n_total if n_total > 0 else 0
            else:
                top_category = np.nan
                top_category_share = np.nan
            
            results.append({
                'domain': domain,
                'feature': feature,
                'type': 'categorical',
                'n_nonmissing': n_nonmissing,
                'n_distinct': n_distinct,
                'top_category': top_category,
                'top_category_share': top_category_share
            })
    
    return pd.DataFrame(results)


def create_headline_features_by_year_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create headline features by year sheet (boolean features only).
    
    Args:
        df: Full DataFrame
        
    Returns:
        DataFrame with headline features by year
    """
    if "ingangsdatum" not in df.columns:
        return pd.DataFrame(columns=['start_year', 'domain', 'feature', 'share_true', 'n_rows_year'])
    
    df_copy = df.copy()
    df_copy["ingangsdatum"] = parse_cao_date_series(df_copy["ingangsdatum"], dayfirst=True)
    df_copy["start_year"] = df_copy["ingangsdatum"].dt.year
    
    results = []
    
    for year, year_group in df_copy.groupby("start_year"):
        if pd.isna(year):
            continue
        
        year = int(year)
        n_rows_year = len(year_group)
        
        # Only boolean features
        for domain, features in HEADLINE_FEATURES.items():
            for feature in features:
                if feature not in year_group.columns:
                    continue
                
                bool_series = normalize_boolean(year_group[feature])
                valid_mask = bool_series.notna()
                if valid_mask.sum() > 0:
                    share_true = (bool_series[valid_mask] == True).mean()
                else:
                    share_true = np.nan
                
                results.append({
                    'start_year': year,
                    'domain': domain,
                    'feature': feature,
                    'share_true': share_true,
                    'n_rows_year': n_rows_year
                })
    
    return pd.DataFrame(results)


def create_numeric_overall_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create numeric overall statistics sheet.
    
    Args:
        df: Full DataFrame
        
    Returns:
        DataFrame with numeric variable statistics
    """
    results = []
    
    for var in NUMERIC_VARS:
        if var not in df.columns:
            continue
        
        stats = describe_numeric(df[var])
        row = {
            'variable': var,
            'n_nonmissing': stats['n'],
            'mean': stats['mean'],
            'std': stats['std'],
            'min': stats['min'],
            'p25_25th_percentile': stats['p25'],
            'median': stats['median'],
            'p75_75th_percentile': stats['p75'],
            'max': stats['max']
        }
        results.append(row)
    
    return pd.DataFrame(results)


def create_numeric_by_period_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create numeric statistics by period sheet.
    
    Args:
        df: Full DataFrame
        
    Returns:
        DataFrame with numeric statistics by period
    """
    if "ingangsdatum" not in df.columns:
        return pd.DataFrame(columns=['variable', 'period', 'n_nonmissing', 'mean', 'median'])
    
    df_copy = df.copy()
    df_copy["ingangsdatum"] = parse_cao_date_series(df_copy["ingangsdatum"], dayfirst=True)
    df_copy["start_year"] = df_copy["ingangsdatum"].dt.year
    
    # Define periods
    def get_period(year):
        if pd.isna(year):
            return np.nan
        year = int(year)
        if year <= 2009:
            return "≤2009"
        elif 2010 <= year <= 2019:
            return "2010–2019"
        else:
            return "≥2020"
    
    df_copy["period"] = df_copy["start_year"].apply(get_period)
    
    results = []
    
    for var in NUMERIC_VARS:
        if var not in df.columns:
            continue
        
        for period in ["≤2009", "2010–2019", "≥2020"]:
            period_data = df_copy[df_copy["period"] == period][var]
            numeric_data = pd.to_numeric(period_data, errors='coerce')
            non_null = numeric_data.dropna()
            
            if len(non_null) > 0:
                results.append({
                    'variable': var,
                    'period': period,
                    'n_nonmissing': len(non_null),
                    'mean': non_null.mean(),
                    'median': non_null.median()
                })
            else:
                results.append({
                    'variable': var,
                    'period': period,
                    'n_nonmissing': 0,
                    'mean': np.nan,
                    'median': np.nan
                })
    
    return pd.DataFrame(results)


def create_modern_before_after_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create modern policy before/after comparison sheet.
    
    Args:
        df: Full DataFrame
        
    Returns:
        DataFrame with early vs late period comparisons
    """
    if "ingangsdatum" not in df.columns:
        return pd.DataFrame(columns=['feature', 'share_true_early_period_year_lt_2010', 
                                     'share_true_late_period_year_ge_2020', 'diff_late_minus_early'])
    
    df_copy = df.copy()
    df_copy["ingangsdatum"] = parse_cao_date_series(df_copy["ingangsdatum"], dayfirst=True)
    df_copy["start_year"] = df_copy["ingangsdatum"].dt.year
    
    # Define periods
    early_mask = df_copy["start_year"] < 2010
    late_mask = df_copy["start_year"] >= 2020
    
    results = []
    
    for feature in MODERN_FEATURES:
        if feature not in df.columns:
            continue
        
        bool_series = normalize_boolean(df_copy[feature])
        
        # Early period
        early_data = bool_series[early_mask]
        early_valid = early_data.dropna()
        share_true_early = (early_valid == True).mean() if len(early_valid) > 0 else np.nan
        
        # Late period
        late_data = bool_series[late_mask]
        late_valid = late_data.dropna()
        share_true_late = (late_valid == True).mean() if len(late_valid) > 0 else np.nan
        
        # Difference
        if not pd.isna(share_true_early) and not pd.isna(share_true_late):
            diff = share_true_late - share_true_early
        else:
            diff = np.nan
        
        results.append({
            'feature': feature,
            'share_true_early_period_year_lt_2010': share_true_early,
            'share_true_late_period_year_ge_2020': share_true_late,
            'diff_late_minus_early': diff
        })
    
    return pd.DataFrame(results)


def create_missing_dates_overview_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create overview sheet for missing date analysis.
    
    Args:
        df: Input DataFrame
        
    Returns:
        DataFrame with missing date overview statistics
    """
    results = []
    total_rows = len(df)
    
    # Parse date columns
    df_copy = df.copy()
    df_copy["ingangsdatum"] = parse_cao_date_series(df_copy["ingangsdatum"], dayfirst=True)
    
    # Try to find contract start date column
    contract_date_col = None
    if "general_start_date" in df_copy.columns:
        contract_date_col = "general_start_date"
    elif "start_date" in df_copy.columns:
        contract_date_col = "start_date"
    else:
        return pd.DataFrame(columns=['metric', 'value'])
    
    df_copy[contract_date_col] = pd.to_datetime(df_copy[contract_date_col], errors='coerce')
    
    # Basic counts
    missing_ingangsdatum = df_copy["ingangsdatum"].isna().sum()
    missing_contract_start = df_copy[contract_date_col].isna().sum()
    missing_both = (df_copy["ingangsdatum"].isna() & df_copy[contract_date_col].isna()).sum()
    missing_only_ingangsdatum = (df_copy["ingangsdatum"].isna() & df_copy[contract_date_col].notna()).sum()
    missing_only_contract_start = (df_copy["ingangsdatum"].notna() & df_copy[contract_date_col].isna()).sum()
    have_both = (df_copy["ingangsdatum"].notna() & df_copy[contract_date_col].notna()).sum()
    
    results.append({'metric': 'Total rows', 'value': total_rows})
    results.append({'metric': f'Rows missing ingangsdatum', 'value': missing_ingangsdatum})
    results.append({'metric': f'Rows missing {contract_date_col}', 'value': missing_contract_start})
    results.append({'metric': 'Rows missing both dates', 'value': missing_both})
    results.append({'metric': 'Rows missing only ingangsdatum', 'value': missing_only_ingangsdatum})
    results.append({'metric': f'Rows missing only {contract_date_col}', 'value': missing_only_contract_start})
    results.append({'metric': 'Rows with both dates', 'value': have_both})
    
    # CAO-level analysis
    if "cao_number" in df_copy.columns:
        missing_ing_caos = df_copy[df_copy["ingangsdatum"].isna()]["cao_number"].nunique()
        missing_contract_caos = df_copy[df_copy[contract_date_col].isna()]["cao_number"].nunique()
        missing_both_caos = df_copy[(df_copy["ingangsdatum"].isna() & df_copy[contract_date_col].isna())]["cao_number"].nunique()
        
        results.append({'metric': 'CAOs missing ingangsdatum', 'value': missing_ing_caos})
        results.append({'metric': f'CAOs missing {contract_date_col}', 'value': missing_contract_caos})
        results.append({'metric': 'CAOs missing both dates', 'value': missing_both_caos})
    
    # Check if disjoint
    if missing_both == 0:
        results.append({'metric': 'Sets disjoint?', 'value': 'Yes (no overlap)'})
    else:
        results.append({'metric': 'Sets disjoint?', 'value': f'No (overlap: {missing_both} rows)'})
    
    return pd.DataFrame(results)


def create_missing_dates_details_sheet(df: pd.DataFrame, missing_type: str) -> pd.DataFrame:
    """
    Create detailed sheet for rows missing specific date.
    
    Args:
        df: Input DataFrame
        missing_type: 'ingangsdatum' or 'contract_start'
        
    Returns:
        DataFrame with details of missing date rows
    """
    df_copy = df.copy()
    df_copy["ingangsdatum"] = parse_cao_date_series(df_copy["ingangsdatum"], dayfirst=True)
    
    # Try to find contract start date column
    contract_date_col = None
    if "general_start_date" in df_copy.columns:
        contract_date_col = "general_start_date"
    elif "start_date" in df_copy.columns:
        contract_date_col = "start_date"
    else:
        return pd.DataFrame()
    
    df_copy[contract_date_col] = pd.to_datetime(df_copy[contract_date_col], errors='coerce')
    
    # Filter to missing rows
    if missing_type == 'ingangsdatum':
        missing_df = df_copy[df_copy["ingangsdatum"].isna()].copy()
    elif missing_type == 'contract_start':
        missing_df = df_copy[df_copy[contract_date_col].isna()].copy()
    else:
        return pd.DataFrame()
    
    if len(missing_df) == 0:
        return pd.DataFrame(columns=['cao_number', 'file_name', 'has_other_date', 'other_date_value'])
    
    # Select key columns for display
    key_cols = []
    if "cao_number" in missing_df.columns:
        key_cols.append("cao_number")
    if "file_name" in missing_df.columns:
        key_cols.append("file_name")
    if "id" in missing_df.columns:
        key_cols.append("id")
    
    # Add other date information (format dates as DD/MM/YYYY for Excel)
    if missing_type == 'ingangsdatum':
        missing_df['has_other_date'] = missing_df[contract_date_col].notna()
        missing_df['other_date_value'] = missing_df[contract_date_col].apply(
            lambda x: x.strftime('%d/%m/%Y') if pd.notna(x) else ''
        )
    else:
        missing_df['has_other_date'] = missing_df["ingangsdatum"].notna()
        missing_df['other_date_value'] = missing_df["ingangsdatum"].apply(
            lambda x: x.strftime('%d/%m/%Y') if pd.notna(x) else ''
        )
    
    # Add other date columns
    date_cols = [col for col in missing_df.columns if 'date' in col.lower() or 'datum' in col.lower()]
    for col in date_cols[:5]:  # Limit to first 5 date columns
        if col not in key_cols and col not in ['has_other_date', 'other_date_value']:
            key_cols.append(col)
    
    result_cols = key_cols + ['has_other_date', 'other_date_value']
    result_cols = [col for col in result_cols if col in missing_df.columns]
    
    result_df = missing_df[result_cols].copy()
    
    # Format datetime columns as DD/MM/YYYY strings so Excel shows date without time
    for col in result_df.columns:
        if pd.api.types.is_datetime64_any_dtype(result_df[col]):
            result_df[col] = result_df[col].apply(
                lambda x: x.strftime('%d/%m/%Y') if pd.notna(x) else ''
            )
    
    # Replace NaN with None for Excel compatibility
    result_df = result_df.where(pd.notna(result_df), None)
    
    return result_df


def create_missing_dates_variable_fill_sheet(df: pd.DataFrame, missing_type: str) -> pd.DataFrame:
    """
    Create sheet showing variable fill status for rows missing dates.
    
    Args:
        df: Input DataFrame
        missing_type: 'ingangsdatum', 'contract_start', or 'both'
        
    Returns:
        DataFrame with variable fill statistics
    """
    df_copy = df.copy()
    df_copy["ingangsdatum"] = parse_cao_date_series(df_copy["ingangsdatum"], dayfirst=True)
    
    # Try to find contract start date column
    contract_date_col = None
    if "general_start_date" in df_copy.columns:
        contract_date_col = "general_start_date"
    elif "start_date" in df_copy.columns:
        contract_date_col = "start_date"
    else:
        return pd.DataFrame()
    
    df_copy[contract_date_col] = pd.to_datetime(df_copy[contract_date_col], errors='coerce')
    
    # Filter to missing rows
    if missing_type == 'ingangsdatum':
        missing_df = df_copy[df_copy["ingangsdatum"].isna()].copy()
        group_name = "Missing ingangsdatum"
    elif missing_type == 'contract_start':
        missing_df = df_copy[df_copy[contract_date_col].isna()].copy()
        group_name = f"Missing {contract_date_col}"
    elif missing_type == 'both':
        missing_df = df_copy[(df_copy["ingangsdatum"].isna() & df_copy[contract_date_col].isna())].copy()
        group_name = "Missing both dates"
    else:
        return pd.DataFrame()
    
    if len(missing_df) == 0:
        return pd.DataFrame(columns=['variable_name', 'n_nonmissing', 'share_nonmissing', 'n_missing'])
    
    total_rows = len(missing_df)
    results = []
    
    # Analyze key variable groups
    variable_groups = {
        'general_': [col for col in df_copy.columns if col.startswith('general_')],
        'bonus_': [col for col in df_copy.columns if col.startswith('bonus_')],
        'pension_': [col for col in df_copy.columns if col.startswith('pension_')],
        'leave_': [col for col in df_copy.columns if col.startswith('leave_')],
        'term_': [col for col in df_copy.columns if col.startswith('term_')],
        'overtime_': [col for col in df_copy.columns if col.startswith('overtime_')],
        'training_': [col for col in df_copy.columns if col.startswith('training_')],
        'homeoffice_': [col for col in df_copy.columns if col.startswith('homeoffice_')],
        'contract_': [col for col in df_copy.columns if col.startswith('contract_')],
        'fringe_': [col for col in df_copy.columns if col.startswith('fringe_')],
        'safety_': [col for col in df_copy.columns if col.startswith('safety_')],
        'childcare_': [col for col in df_copy.columns if col.startswith('childcare_')],
        'ai_': [col for col in df_copy.columns if col.startswith('ai_')],
    }
    
    # Analyze domain flags
    domain_flags = {
        "bonus": "bonus_has_bonus_schemes",
        "pension": "pension_has_pension_scheme",
        "leave": "leave_has_leave_enhancements",
        "termination": "term_has_termination_rules",
        "overtime": "overtime_has_overtime_rules",
        "training": "training_has_training_rights",
        "homeoffice": "homeoffice_has_homeoffice_rights",
        "contract": "contract_has_contract_type_rules",
        "fringe": "fringe_has_fringe_benefits",
    }
    
    # Domain presence
    for domain, flag_col in domain_flags.items():
        if flag_col in missing_df.columns:
            bool_series = normalize_boolean(missing_df[flag_col])
            n_true = (bool_series == True).sum()
            n_nonmissing = bool_series.notna().sum()
            results.append({
                'variable_name': f'{domain}_domain_present',
                'n_nonmissing': n_nonmissing,
                'share_nonmissing': n_nonmissing / total_rows if total_rows > 0 else 0,
                'n_missing': total_rows - n_nonmissing,
                'n_true': n_true,
                'share_true': n_true / total_rows if total_rows > 0 else 0
            })
    
    # Sample of key variables from each domain (limit to avoid too many rows)
    key_vars = []
    for prefix, cols in variable_groups.items():
        # Get boolean and key numeric variables
        bool_cols = [c for c in cols if c in missing_df.columns and 
                     (c.endswith('_present') or c.endswith('_applies') or 
                      missing_df[c].dtype == bool)]
        key_vars.extend(bool_cols[:3])  # Limit to 3 per domain
    
    for var in key_vars:
        if var in missing_df.columns:
            series = missing_df[var]
            n_nonmissing = series.notna().sum()
            var_type = infer_var_type(series, var)
            
            row = {
                'variable_name': var,
                'n_nonmissing': n_nonmissing,
                'share_nonmissing': n_nonmissing / total_rows if total_rows > 0 else 0,
                'n_missing': total_rows - n_nonmissing,
                'variable_type': var_type
            }
            
            if var_type == "boolean":
                bool_series = normalize_boolean(series)
                n_true = (bool_series == True).sum()
                row['n_true'] = n_true
                row['share_true'] = n_true / total_rows if total_rows > 0 else 0
            
            results.append(row)
    
    result_df = pd.DataFrame(results)
    
    # Replace NaN with None for Excel compatibility
    result_df = result_df.where(pd.notna(result_df), None)
    
    return result_df


def create_cao_dates_timeline_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create sheet with ordered list of file dates per CAO number with counts.
    
    For each CAO number, groups files by unique ingangsdatum date and shows
    the date with the count of files for that date, in chronological order.
    Format: CAO number, n_files (total), then file_1_date, file_1_count, 
    file_2_date, file_2_count, etc.
    
    Args:
        df: Input DataFrame with cao_number and ingangsdatum columns
        
    Returns:
        DataFrame with CAO numbers, total file counts, and date-count pairs
    """
    if "cao_number" not in df.columns or "ingangsdatum" not in df.columns:
        print(f"  Warning: Missing required columns. Available columns: {list(df.columns[:10])}")
        return pd.DataFrame(columns=['cao_number', 'n_files'])
    
    df_copy = df.copy()
    
    # Parse date column (CAO metadata dates are in DD/MM/YYYY format)
    df_copy["ingangsdatum"] = parse_cao_date_series(df_copy["ingangsdatum"], dayfirst=True)
    
    # All unique CAOs from full df (include CAOs with no valid ingangsdatum)
    all_cao_numbers = df_copy["cao_number"].dropna().astype(str).unique()
    # Rows with valid CAO number and date (for timeline content)
    valid_mask = df_copy["cao_number"].notna() & df_copy["ingangsdatum"].notna()
    df_valid = df_copy[valid_mask].copy()
    
    # Convert cao_number to string for consistent grouping
    df_valid["cao_number"] = df_valid["cao_number"].astype(str)
    # File count per CAO from full df (for CAOs with no dates)
    file_count_per_cao = df_copy[df_copy["cao_number"].notna()].groupby(
        df_copy["cao_number"].astype(str)
    ).size()
    
    # Group by CAO number and collect dates (only from df_valid)
    results = []
    max_files = 0
    caos_with_dates = set()
    
    for cao_number, group in df_valid.groupby("cao_number"):
        # Count occurrences of each unique date
        dates = group["ingangsdatum"].dropna()
        if len(dates) == 0:
            continue
        
        caos_with_dates.add(str(cao_number))
        # Count files per unique date
        date_counts = dates.value_counts().sort_index()  # Sort by date
        
        # Format dates as DD.MM.YYYY and create date-count pairs
        date_count_pairs = []
        total_files = 0
        for date, count in date_counts.items():
            try:
                date_str = date.strftime('%d.%m.%Y')
                date_count_pairs.append((date_str, int(count)))
                total_files += count
            except (AttributeError, ValueError):
                continue
        
        if len(date_count_pairs) == 0:
            continue
        
        max_files = max(max_files, len(date_count_pairs))
        
        row = {
            'cao_number': str(cao_number),
            'n_files': total_files
        }
        for i, (date_str, count) in enumerate(date_count_pairs, start=1):
            row[f'file_{i}_date'] = date_str
            row[f'file_{i}_count'] = count
        results.append(row)
    
    # Include CAOs with no valid ingangsdatum (one row each: cao_number, n_files, no dates)
    for cao_number in all_cao_numbers:
        if cao_number in caos_with_dates:
            continue
        n_files = int(file_count_per_cao.get(cao_number, 0))
        if n_files == 0:
            continue
        results.append({
            'cao_number': str(cao_number),
            'n_files': n_files
        })
    
    if len(results) == 0:
        print(f"  Warning: No results generated")
        return pd.DataFrame(columns=['cao_number', 'n_files'])
    
    # Create DataFrame
    result_df = pd.DataFrame(results)
    
    # Ensure all date and count columns exist (fill missing with empty string or 0)
    for i in range(1, max_files + 1):
        date_col = f'file_{i}_date'
        count_col = f'file_{i}_count'
        if date_col not in result_df.columns:
            result_df[date_col] = ''
        if count_col not in result_df.columns:
            result_df[count_col] = 0
        result_df[date_col] = result_df[date_col].fillna('')
        result_df[count_col] = result_df[count_col].fillna(0)
    
    # Reorder columns: cao_number, n_files, then file_1_date, file_1_count, file_2_date, file_2_count, etc.
    if max_files > 0:
        # Create list of column pairs: date, count, date, count, ...
        date_count_cols = []
        for i in range(1, max_files + 1):
            date_col = f'file_{i}_date'
            count_col = f'file_{i}_count'
            if date_col in result_df.columns:
                date_count_cols.append(date_col)
            if count_col in result_df.columns:
                date_count_cols.append(count_col)
        
        # Only reorder if all columns exist
        available_cols = ['cao_number', 'n_files'] + date_count_cols
        result_df = result_df[available_cols]
    else:
        # If no files, just ensure basic columns
        result_df = result_df[['cao_number', 'n_files']]
    
    # Sort by CAO number
    result_df = result_df.sort_values('cao_number')
    
    print(f"  Created timeline sheet with {len(result_df)} CAOs, max {max_files} unique dates per CAO")
    
    return result_df


# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    """Main entry point for descriptives script."""
    print("="*80)
    print("CAO Non-Salary Descriptives Script")
    print("="*80)
    
    # Load data
    print(f"\nLoading data from: {INPUT_CSV_PATH}")
    try:
        df = pd.read_csv(INPUT_CSV_PATH, sep=';', encoding='utf-8')
        print(f"  Loaded {len(df)} rows and {len(df.columns)} columns")
    except Exception as e:
        print(f"  ERROR: Could not load input file: {e}")
        return
    
    if len(df) == 0:
        print("  ERROR: Input file is empty")
        return
    
    # Parse date column (CAO metadata dates are in DD/MM/YYYY format)
    if "ingangsdatum" in df.columns:
        df["ingangsdatum"] = parse_cao_date_series(df["ingangsdatum"], dayfirst=True)
        print(f"  Parsed ingangsdatum as datetime")
    else:
        print(f"  Warning: ingangsdatum column not found")
    
    # Build latest CAO view
    print("\nBuilding latest CAO view...")
    try:
        df_latest = build_latest_cao_view(df)
        print(f"  Latest CAO view: {len(df_latest)} unique CAOs")
    except Exception as e:
        print(f"  Warning: Error building latest CAO view: {e}")
        df_latest = pd.DataFrame()
    
    # Create output directory
    output_path = Path(OUTPUT_EXCEL_PATH)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Generate all sheets
    print("\nGenerating descriptive statistics sheets...")
    
    sheets = {}
    
    try:
        print("  Creating sheet: 00_variable_health")
        sheets["00_variable_health"] = create_variable_health_sheet(df)
    except Exception as e:
        print(f"  Warning: Error creating variable_health sheet: {e}")
        sheets["00_variable_health"] = pd.DataFrame()
    
    try:
        print("  Creating sheet: 00b_unused_fields")
        sheets["00b_unused_fields"] = create_unused_fields_sheet(df)
    except Exception as e:
        print(f"  Warning: Error creating unused_fields sheet: {e}")
        sheets["00b_unused_fields"] = pd.DataFrame()
    
    try:
        print("  Creating sheet: 01_sample_overview")
        sheets["01_sample_overview"] = create_sample_overview_sheet(df, df_latest)
    except Exception as e:
        print(f"  Warning: Error creating sample_overview sheet: {e}")
        sheets["01_sample_overview"] = pd.DataFrame()
    
    try:
        print("  Creating sheet: 01a_cao_dates_timeline")
        sheets["01a_cao_dates_timeline"] = create_cao_dates_timeline_sheet(df)
    except Exception as e:
        print(f"  Warning: Error creating cao_dates_timeline sheet: {e}")
        sheets["01a_cao_dates_timeline"] = pd.DataFrame()
    
    try:
        print("  Creating sheet: 02_domain_coverage_latest")
        sheets["02_domain_coverage_latest"] = create_domain_coverage_latest_sheet(df_latest)
    except Exception as e:
        print(f"  Warning: Error creating domain_coverage_latest sheet: {e}")
        sheets["02_domain_coverage_latest"] = pd.DataFrame()
    
    try:
        print("  Creating sheet: 03_domain_coverage_by_year")
        sheets["03_domain_coverage_by_year"] = create_domain_coverage_by_year_sheet(df)
    except Exception as e:
        print(f"  Warning: Error creating domain_coverage_by_year sheet: {e}")
        sheets["03_domain_coverage_by_year"] = pd.DataFrame()
    
    try:
        print("  Creating sheet: 04_headline_features_latest")
        sheets["04_headline_features_latest"] = create_headline_features_latest_sheet(df_latest)
    except Exception as e:
        print(f"  Warning: Error creating headline_features_latest sheet: {e}")
        sheets["04_headline_features_latest"] = pd.DataFrame()
    
    try:
        print("  Creating sheet: 05_headline_features_by_year")
        sheets["05_headline_features_by_year"] = create_headline_features_by_year_sheet(df)
    except Exception as e:
        print(f"  Warning: Error creating headline_features_by_year sheet: {e}")
        sheets["05_headline_features_by_year"] = pd.DataFrame()
    
    try:
        print("  Creating sheet: 06_numeric_overall")
        sheets["06_numeric_overall"] = create_numeric_overall_sheet(df)
    except Exception as e:
        print(f"  Warning: Error creating numeric_overall sheet: {e}")
        sheets["06_numeric_overall"] = pd.DataFrame()
    
    try:
        print("  Creating sheet: 07_numeric_by_period")
        sheets["07_numeric_by_period"] = create_numeric_by_period_sheet(df)
    except Exception as e:
        print(f"  Warning: Error creating numeric_by_period sheet: {e}")
        sheets["07_numeric_by_period"] = pd.DataFrame()
    
    try:
        print("  Creating sheet: 08_modern_before_after")
        sheets["08_modern_before_after"] = create_modern_before_after_sheet(df)
    except Exception as e:
        print(f"  Warning: Error creating modern_before_after sheet: {e}")
        sheets["08_modern_before_after"] = pd.DataFrame()
    
    # Missing dates analysis sheets
    try:
        print("  Creating sheet: 09_missing_dates_overview")
        sheets["09_missing_dates_overview"] = create_missing_dates_overview_sheet(df)
    except Exception as e:
        print(f"  Warning: Error creating missing_dates_overview sheet: {e}")
        sheets["09_missing_dates_overview"] = pd.DataFrame()
    
    try:
        print("  Creating sheet: 10_missing_ingangsdatum_details")
        sheets["10_missing_ingangsdatum_details"] = create_missing_dates_details_sheet(df, 'ingangsdatum')
    except Exception as e:
        print(f"  Warning: Error creating missing_ingangsdatum_details sheet: {e}")
        sheets["10_missing_ingangsdatum_details"] = pd.DataFrame()
    
    try:
        print("  Creating sheet: 11_missing_contract_start_det")
        sheets["11_missing_contract_start_det"] = create_missing_dates_details_sheet(df, 'contract_start')
    except Exception as e:
        print(f"  Warning: Error creating missing_contract_start_details sheet: {e}")
        sheets["11_missing_contract_start_det"] = pd.DataFrame()
    
    try:
        print("  Creating sheet: 12_missing_ingangs_var_fill")
        sheets["12_missing_ingangs_var_fill"] = create_missing_dates_variable_fill_sheet(df, 'ingangsdatum')
    except Exception as e:
        print(f"  Warning: Error creating missing_ingangsdatum_variable_fill sheet: {e}")
        sheets["12_missing_ingangs_var_fill"] = pd.DataFrame()
    
    try:
        print("  Creating sheet: 13_missing_contract_var_fill")
        sheets["13_missing_contract_var_fill"] = create_missing_dates_variable_fill_sheet(df, 'contract_start')
    except Exception as e:
        print(f"  Warning: Error creating missing_contract_start_variable_fill sheet: {e}")
        sheets["13_missing_contract_var_fill"] = pd.DataFrame()
    
    try:
        print("  Creating sheet: 14_missing_both_var_fill")
        sheets["14_missing_both_var_fill"] = create_missing_dates_variable_fill_sheet(df, 'both')
    except Exception as e:
        print(f"  Warning: Error creating missing_both_dates_variable_fill sheet: {e}")
        sheets["14_missing_both_var_fill"] = pd.DataFrame()
    
    # Write to Excel with explanatory notes
    print(f"\nWriting to Excel: {OUTPUT_EXCEL_PATH}")
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        
        # First write all dataframes
        with pd.ExcelWriter(OUTPUT_EXCEL_PATH, engine='openpyxl') as writer:
            for sheet_name, sheet_df in sheets.items():
                if len(sheet_df) > 0:
                    # Replace NaN with None for Excel compatibility
                    sheet_df_clean = sheet_df.where(pd.notna(sheet_df), None)
                    # Convert any remaining problematic types
                    for col in sheet_df_clean.columns:
                        if sheet_df_clean[col].dtype == 'object':
                            # Replace any remaining NaN/NaT with empty string
                            sheet_df_clean[col] = sheet_df_clean[col].fillna('')
                    sheet_df_clean.to_excel(writer, sheet_name=sheet_name, index=False)
                    print(f"  Wrote sheet '{sheet_name}' with {len(sheet_df_clean)} rows")
                else:
                    # Write empty sheet with headers if possible
                    sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    print(f"  Wrote empty sheet '{sheet_name}'")
        
        # Now add explanatory notes using openpyxl
        wb = load_workbook(OUTPUT_EXCEL_PATH)
        
        # Add notes to specific sheets
        notes = {
            "08_modern_before_after": [
                "NOTES:",
                "Period definitions:",
                "  - Early period: Contracts with start year < 2010",
                "  - Late period: Contracts with start year >= 2020",
                "",
                "share_true_early_period_year_lt_2010: Share of contracts in early period with feature = True",
                "share_true_late_period_year_ge_2020: Share of contracts in late period with feature = True",
                "diff_late_minus_early: Difference (late - early), positive values indicate increase over time",
                "",
                "Note: Contracts from 2010-2019 are excluded from this comparison."
            ],
            "01_sample_overview": [
                "NOTES:",
                "This sheet contains five sections:",
                "",
                "1. Panel by start year (section='panel_by_year'):",
                "   - n_contracts: Number of contract episodes starting in that year",
                "   - n_cao: Number of unique CAOs with contracts starting in that year",
                "   - share_avv: Share of contracts in that year with AVV (Algemeen Verbindend Verklaring) applies",
                "",
                "2. Cross-section latest CAO (section='cross_section_latest'):",
                "   - One row per unique CAO number, using the contract with the latest start date",
                "   - n_cao: Total number of unique CAOs in latest view",
                "   - share_[scope_type]: Share of CAOs with each scope type (sectoral, single_company, etc.)",
                "   - n_[scope_type]: Count of CAOs with each scope type",
                "   - n_sbi_codes: Number of unique SBI codes",
                "",
                "3. CAO scope type distribution (section='scope_type_distribution'):",
                "   - One row per scope type category",
                "   - scope_type: The scope type category (sectoral, single_company, group, etc.)",
                "   - n_cao: Number of contracts with this scope type",
                "   - share: Share of all contracts with this scope type",
                "",
                "4. Retroactivity summary (section='retroactivity'):",
                "   - share_retroactive: Share of all contracts with retroactive application",
                "   - share_backpay_due: Share of retroactive contracts requiring backpay (conditional on retroactive=True)",
                "   - retro_length_median/p25/p75: Median and quartiles of retroactivity period length in days",
                "   - share_interest_or_surcharge: Share of retroactive contracts with interest/surcharge on backpay",
                "",
                "5. Date comparison (section='date_comparison'):",
                "   - Compares ingangsdatum (date from website) vs general_start_date (date from PDF)",
                "   - date_diff_days: PDF date - website date (positive = PDF later, negative = PDF earlier)",
                "   - n_exact_match: Number of contracts where dates match exactly",
                "   - n_pdf_later_than_website: PDF date is after website date",
                "   - n_pdf_earlier_than_website: PDF date is before website date",
                "   - Large discrepancy: Absolute difference > 30 days"
            ],
            "01a_cao_dates_timeline": [
                "NOTES:",
                "Timeline of file dates per CAO number with counts, ordered chronologically.",
                "",
                "For each CAO number, this sheet groups files by unique ingangsdatum date",
                "and shows the date with the count of files for that date, in chronological order.",
                "Dates are formatted as DD.MM.YYYY (e.g., 03.09.2010).",
                "",
                "Columns:",
                "  - cao_number: The CAO number",
                "  - n_files: Total number of files (sum of all file_x_count values)",
                "  - file_1_date: First unique date (earliest date)",
                "  - file_1_count: Number of files with file_1_date",
                "  - file_2_date: Second unique date",
                "  - file_2_count: Number of files with file_2_date",
                "  - ... (additional date-count pairs as needed)",
                "",
                "Example:",
                "  CAO number 10 might have:",
                "    - file_1_date: 03.09.2010, file_1_count: 3",
                "    - file_2_date: 03.11.2020, file_2_count: 2",
                "    - file_3_date: 01.01.2025, file_3_count: 1",
                "    - n_files: 6 (sum of all counts)"
            ],
            "02_domain_coverage_latest": [
                "NOTES:",
                "Domain coverage for the latest CAO view (one row per unique CAO, using contract with latest start date).",
                "",
                "n_cao_with_domain: Number of CAOs (in latest view) that have this domain present",
                "share_cao_with_domain: Share of all CAOs (in latest view) that have this domain present",
                "",
                "For safety and childcare domains, presence is determined by OR logic across multiple columns.",
                "For other domains, presence is determined by a single boolean column."
            ],
            "03_domain_coverage_by_year": [
                "NOTES:",
                "Domain coverage by contract start year.",
                "",
                "share_with_domain: Share of contracts starting in that year with the domain present",
                "n_rows_year: Total number of contract episodes starting in that year",
                "",
                "For safety and childcare domains, presence is determined by OR logic across multiple columns.",
                "For other domains, presence is determined by a single boolean column."
            ],
            "04_headline_features_latest": [
                "NOTES:",
                "Headline policy features for the latest CAO view (one row per unique CAO, using contract with latest start date).",
                "",
                "For boolean features:",
                "  - share_true: Share of CAOs with feature = True",
                "  - n_true/n_false/n_missing: Counts of True/False/Missing values",
                "",
                "For categorical features:",
                "  - n_nonmissing: Number of CAOs with non-missing values",
                "  - n_distinct: Number of distinct categories",
                "  - top_category: Most frequent category",
                "  - top_category_share: Share of CAOs with the top category"
            ],
            "05_headline_features_by_year": [
                "NOTES:",
                "Headline policy features by contract start year (boolean features only).",
                "",
                "share_true: Share of contracts starting in that year with feature = True",
                "n_rows_year: Total number of contract episodes starting in that year"
            ],
            "00b_unused_fields": [
                "NOTES:",
                "This sheet lists all fields that are completely unused in the dataset.",
                "",
                "For non-boolean fields: n_nonmissing = 0 (all values are missing)",
                "For boolean fields: n_true = 0 (all values are False or missing, never True)",
                "",
                "inferred_type: Type of the field (boolean, numeric, categorical, date, text, id/other)",
                "n_nonmissing: Number of non-missing values (should be 0 for all rows)",
                "n_true/n_false: For boolean fields, counts of True/False values",
                "n_missing: Number of missing values"
            ],
            "06_numeric_overall": [
                "NOTES:",
                "Overall descriptive statistics for numeric variables across all contracts.",
                "",
                "n_nonmissing: Number of non-missing values",
                "p25_25th_percentile: 25th percentile (first quartile)",
                "p75_75th_percentile: 75th percentile (third quartile)",
                "All statistics calculated only on non-missing values."
            ],
            "07_numeric_by_period": [
                "NOTES:",
                "Period definitions:",
                "  - ≤2009: Contracts with start year <= 2009",
                "  - 2010–2019: Contracts with start year between 2010 and 2019 (inclusive)",
                "  - ≥2020: Contracts with start year >= 2020",
                "",
                "n_nonmissing: Number of non-missing values within the period",
                "Statistics are calculated only for non-missing values within each period."
            ],
            "09_missing_dates_overview": [
                "NOTES:",
                "Overview of missing date analysis for ingangsdatum and general_start_date.",
                "See the metric/value columns above for current counts.",
                "  - Rows missing ingangsdatum: rows where website date is missing",
                "  - Rows missing general_start_date: rows where PDF date is missing",
                "  - Sets disjoint? Yes if no row is missing both dates."
            ],
            "10_missing_ingangsdatum_details": [
                "NOTES:",
                "Detailed list of rows missing ingangsdatum (website date).",
                "has_other_date: Whether the row has general_start_date",
                "other_date_value: The value of general_start_date if available (DD/MM/YYYY)."
            ],
            "11_missing_contract_start_det": [
                "NOTES:",
                "Detailed list of rows missing general_start_date (PDF date).",
                "has_other_date: Whether the row has ingangsdatum",
                "other_date_value: The value of ingangsdatum if available (DD/MM/YYYY)."
            ],
            "12_missing_ingangs_var_fill": [
                "NOTES:",
                "Variable fill status for rows missing ingangsdatum (46 rows).",
                "",
                "Shows how well variables are filled for this subset:",
                "  - Domain presence flags (bonus, pension, leave, etc.)",
                "  - Key boolean and numeric variables from each domain",
                "",
                "n_nonmissing: Number of rows with non-missing values",
                "share_nonmissing: Proportion of rows with non-missing values",
                "n_true: For boolean variables, number of rows with True",
                "share_true: For boolean variables, proportion of rows with True"
            ],
            "13_missing_contract_var_fill": [
                "NOTES:",
                "Variable fill status for rows missing general_start_date (5 rows).",
                "",
                "Shows how well variables are filled for this subset:",
                "  - Domain presence flags (bonus, pension, leave, etc.)",
                "  - Key boolean and numeric variables from each domain",
                "",
                "n_nonmissing: Number of rows with non-missing values",
                "share_nonmissing: Proportion of rows with non-missing values",
                "n_true: For boolean variables, number of rows with True",
                "share_true: For boolean variables, proportion of rows with True"
            ],
            "14_missing_both_var_fill": [
                "NOTES:",
                "Variable fill status for rows missing both dates (1 row: CAO 1393).",
                "",
                "This single row has:",
                "  - expiratiedatum and datum_kennisgeving",
                "  - Domain information (bonus, pension, leave, termination, etc.)",
                "  - Missing both start dates (ingangsdatum and general_start_date)",
                "",
                "n_nonmissing: Number of rows with non-missing values (should be 0 or 1)",
                "share_nonmissing: Proportion of rows with non-missing values",
                "n_true: For boolean variables, number of rows with True",
                "share_true: For boolean variables, proportion of rows with True"
            ]
        }
        
        # Add notes to sheets
        for sheet_name, note_lines in notes.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Find the last row with data
                max_row = ws.max_row
                # Start notes 2 rows below the data
                note_start_row = max_row + 3
                
                for i, note_line in enumerate(note_lines):
                    cell = ws.cell(row=note_start_row + i, column=1)
                    cell.value = note_line
                    # Make notes italic
                    from openpyxl.styles import Font
                    try:
                        existing_font = cell.font
                        cell.font = Font(italic=True, name=existing_font.name if existing_font else 'Calibri', 
                                       size=existing_font.size if existing_font else 11, 
                                       bold=existing_font.bold if existing_font else False, 
                                       color=existing_font.color if existing_font and existing_font.color else None)
                    except Exception:
                        # Fallback if font access fails
                        cell.font = Font(italic=True)
        
        wb.save(OUTPUT_EXCEL_PATH)
        print(f"\n✓ Successfully created Excel workbook with {len(sheets)} sheets and explanatory notes")
        print(f"  Output: {OUTPUT_EXCEL_PATH}")
    except Exception as e:
        print(f"\n  ERROR: Could not write Excel file: {e}")
        import traceback
        traceback.print_exc()
        return
    
    print("\n" + "="*80)
    print("Script completed successfully!")
    print("="*80)


if __name__ == "__main__":
    main()

