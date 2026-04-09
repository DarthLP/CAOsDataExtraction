"""
CAO Salary Descriptives Script

This script reads the salary CSV output and produces a multi-sheet Excel workbook
with comprehensive descriptive statistics including variable health, sample overview,
salary slots coverage, worker profiles, amounts by unit, increase percentages, level trends,
and row slot counts.

USAGE:
    python scripts/excel_analysis/descriptives_salary.py

INPUT:
    - outputs/excel/new_results/extracted_data_salary.csv

OUTPUT:
    - outputs/analysis/salary_descriptives.xlsx (multi-sheet Excel workbook)
    - outputs/analysis/salary_descriptives_excluded_no_salary.csv (summary + fully no-salary files; sep ';')
    - outputs/analysis/salary_increase_events_derived.csv (always written; sep ';')
    - outputs/analysis/salary_increase_conversion_diagnostics.csv (always written; sep ';')
    - outputs/analysis/salary_increase_csv_vs_diff_comparison.csv (sep ';')
    - outputs/analysis/salary_monthly_band_summary.csv (single-row band exclusion summary; sep ';')

Wide descriptives sheets 00–07 use the analytic sample: wide rows with at least one strictly positive
coerced salary_k_amount. Sheet 08 and the exclusion CSV still describe the full extract (including
cancelled / no-salary rows and files). Derived increase CSV row_id values reference iloc in the input CSV.
"""

import os
import sys
import warnings
from pathlib import Path
from typing import Dict, List, Any, Optional, Union
import pandas as pd
import numpy as np
from datetime import datetime

# Add the parent directory to Python path so we can import utils
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from scripts.excel_analysis.analysis_utils import (
    build_long_salary_from_wide,
    coerce_salary_amount_scalar,
    detect_salary_slot_indices,
    parse_cao_date_series,
)
from scripts.excel_analysis.salary_increase_derivation import (
    compute_band_summary_stats,
    derive_salary_increase_series,
)

# =============================================================================
# PATH CONFIGURATION
# =============================================================================
INPUT_CSV = "outputs/excel/new_results/extracted_data_salary.csv"
OUTPUT_EXCEL = "outputs/analysis/salary_descriptives.xlsx"
OUTPUT_EXCLUSION_CSV = "outputs/analysis/salary_descriptives_excluded_no_salary.csv"

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def normalize_boolean(series: pd.Series) -> pd.Series:
    """
    Normalize boolean values to True/False.
    
    Args:
        series: Series with potentially mixed boolean representations
        
    Returns:
        Series with normalized boolean values
    """
    bool_map = {
        1: True, 0: False,
        "1": True, "0": False,
        "yes": True, "no": False,
        "Yes": True, "No": False,
        "YES": True, "NO": False,
        "true": True, "false": False,
        "True": True, "False": False,
        "TRUE": True, "FALSE": False,
    }

    # Scalar map avoids Series.replace downcasting FutureWarning in modern pandas.
    def _cell(x: Any) -> Any:
        if x is None:
            return np.nan
        try:
            if pd.isna(x):
                return np.nan
        except (ValueError, TypeError):
            pass
        if isinstance(x, (bool, np.bool_)):
            return bool(x)
        if x in bool_map:
            return bool_map[x]
        return x

    result = series.map(_cell)
    
    # Convert to boolean, keeping NaN
    result = result.astype(object)
    result = result.where(result.isin([True, False, np.nan]), np.nan)
    
    return result


def log_memory(label: str, frame: pd.DataFrame) -> None:
    """
    Log approximate DataFrame memory in MB for run diagnostics.

    Args:
        label: Human-readable checkpoint label
        frame: DataFrame to inspect

    Returns:
        None
    """
    try:
        mem_mb = frame.memory_usage(deep=True).sum() / (1024 * 1024)
        print(f"  [MEM] {label}: {mem_mb:,.2f} MB")
    except Exception:
        pass


def infer_var_type(series: pd.Series, name: str) -> str:
    """
    Infer variable type from series and column name.
    
    Args:
        series: Pandas Series
        name: Column name
        
    Returns:
        Type string: "boolean", "numeric", "date", "categorical", "text", "id/other"
    """
    # Check for boolean
    if series.dtype == bool:
        return "boolean"
    
    # Check if values collapse to boolean
    non_null = series.dropna()
    if len(non_null) > 0:
        unique_vals = set(non_null.unique())
        bool_like = {True, False, 1, 0, "1", "0", "yes", "no", "Yes", "No", 
                     "YES", "NO", "true", "false", "True", "False", "TRUE", "FALSE"}
        if unique_vals.issubset(bool_like):
            return "boolean"
    
    # Check for date
    name_low = name.lower()
    if "date" in name_low or "datum" in name_low:
        return "date"
    if pd.api.types.is_datetime64_any_dtype(series):
        return "date"
    
    # Check for text patterns in column name
    if name_low.endswith("_note") or name_low.endswith("_label") or "note" in name_low:
        return "text"
    
    # Try numeric
    numeric_series = pd.to_numeric(series, errors='coerce')
    non_null_numeric = numeric_series.dropna()
    if len(non_null_numeric) > 0:
        # If ≥20% of non-null values are numeric, consider it numeric
        if len(non_null_numeric) >= 0.2 * len(non_null):
            return "numeric"
    
    # Distinguish categorical from text
    n_distinct = non_null.nunique()
    
    # If many distinct values, likely text or id/other
    if n_distinct > 20:
        return "id/other"
    
    # If distinct count is low, likely categorical
    if n_distinct <= 20:
        return "categorical"
    
    # Default
    return "id/other"


def build_long_salary_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Build long format DataFrame from wide salary format.
    
    Args:
        df: Wide format DataFrame with salary_k_* columns
        
    Returns:
        Long format DataFrame with one row per salary point
    """
    identity_cols = [
        "cao_number", "id", "TTW", "ingangsdatum", "expiratiedatum",
        "datum_kennisgeving", "file_name",
        "jobgroup", "step_label", "worker_type", "is_entry", "age_group",
        "education", "ft_hours", "ft_hours_weekly", "permanency", "hours_type",
        "row_note",
    ]
    return build_long_salary_from_wide(df, identity_cols=identity_cols)


# =============================================================================
# SHEET GENERATION FUNCTIONS
# =============================================================================

def create_variable_health_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create variable health sheet with one row per column.
    
    Args:
        df: Input DataFrame
        
    Returns:
        DataFrame with variable health statistics
    """
    total_rows = len(df)
    results = []
    
    for col_name in df.columns:
        series = df[col_name]
        non_null = series.dropna()
        n_nonmissing = len(non_null)
        share_nonmissing = n_nonmissing / total_rows if total_rows > 0 else 0
        n_distinct = non_null.nunique() if len(non_null) > 0 else 0
        
        var_type = infer_var_type(series, col_name)
        
        row = {
            'variable_name': col_name,
            'inferred_type': var_type,
            'n_nonmissing': n_nonmissing,
            'share_nonmissing': share_nonmissing,
            'n_distinct': n_distinct
        }
        
        # Type-specific statistics
        if var_type == "boolean":
            bool_series = normalize_boolean(series)
            n_true = (bool_series == True).sum()
            n_false = (bool_series == False).sum()
            n_missing = total_rows - n_nonmissing
            
            row.update({
                'n_true': n_true,
                'share_true': n_true / total_rows if total_rows > 0 else 0,
                'n_false': n_false,
                'share_false': n_false / total_rows if total_rows > 0 else 0,
                'n_missing': n_missing
            })
        
        elif var_type == "numeric":
            numeric_series = pd.to_numeric(series, errors='coerce')
            non_null_num = numeric_series.dropna()
            
            if len(non_null_num) > 0:
                row.update({
                    'mean': non_null_num.mean(),
                    'median': non_null_num.median(),
                    'min': non_null_num.min(),
                    'max': non_null_num.max(),
                    'p25': non_null_num.quantile(0.25),
                    'p75': non_null_num.quantile(0.75),
                    'share_zero': (non_null_num == 0).sum() / len(non_null_num) if len(non_null_num) > 0 else 0
                })
            else:
                row.update({
                    'mean': np.nan, 'median': np.nan, 'min': np.nan, 'max': np.nan,
                    'p25': np.nan, 'p75': np.nan, 'share_zero': np.nan
                })
        
        elif var_type == "categorical":
            if len(non_null) > 0:
                value_counts = non_null.value_counts()
                top_category = value_counts.index[0] if len(value_counts) > 0 else np.nan
                top_category_count = value_counts.iloc[0] if len(value_counts) > 0 else 0
                row.update({
                    'top_category': top_category,
                    'top_category_share': top_category_count / total_rows if total_rows > 0 else 0
                })
            else:
                row.update({
                    'top_category': np.nan,
                    'top_category_share': np.nan
                })
        
        elif var_type == "date":
            # CAO metadata date columns: use robust parser (DD/MM/YYYY)
            if series.name in ['ingangsdatum', 'expiratiedatum', 'datum_kennisgeving']:
                date_series = parse_cao_date_series(series, dayfirst=True)
            else:
                date_series = pd.to_datetime(series, errors='coerce')
            non_null_dates = date_series.dropna()
            if len(non_null_dates) > 0:
                row.update({
                    'min_date': non_null_dates.min(),
                    'max_date': non_null_dates.max()
                })
            else:
                row.update({
                    'min_date': np.nan,
                    'max_date': np.nan
                })
        
        elif var_type == "text":
            if len(non_null) > 0:
                char_lengths = non_null.astype(str).str.len()
                row.update({
                    'avg_char_length': char_lengths.mean()
                })
            else:
                row.update({
                    'avg_char_length': np.nan
                })
        
        results.append(row)
    
    return pd.DataFrame(results)


def create_sample_overview_sheet(
    df: pd.DataFrame,
    df_long: pd.DataFrame,
    band_summary: Optional[Dict[str, Any]] = None,
) -> pd.DataFrame:
    """
    Create sample overview sheet with multiple blocks.
    
    Args:
        df: Wide format DataFrame
        df_long: Long format DataFrame
        band_summary: Optional dict from salary increase derivation (monthly band exclusions).
        
    Returns:
        DataFrame with sample overview statistics
    """
    results = []
    
    # Block A - Overall dataset structure
    block_a = {}
    block_a['section'] = 'overall_structure'
    block_a['n_rows'] = len(df)  # Analytic wide rows (typically ≥1 positive salary_k_amount per row)
    
    # Count unique files (cao_number + file_name combinations)
    if "cao_number" in df.columns and "file_name" in df.columns:
        unique_files = df[['cao_number', 'file_name']].drop_duplicates()
        block_a['n_unique_files'] = len(unique_files)
    else:
        block_a['n_unique_files'] = np.nan
    
    if "cao_number" in df.columns:
        block_a['n_cao'] = df['cao_number'].nunique()
    else:
        block_a['n_cao'] = np.nan
    
    for col, label in [("jobgroup", "n_jobgroups"), ("step_label", "n_steps"),
                       ("worker_type", "n_worker_types"), ("age_group", "n_age_groups"),
                       ("education", "n_education_categories")]:
        if col in df.columns:
            block_a[label] = df[col].nunique()
        else:
            block_a[label] = np.nan
    
    if "ft_hours_weekly" in df.columns:
        ft_weekly = df["ft_hours_weekly"].dropna()
        block_a['ft_hours_weekly_nonmissing'] = len(ft_weekly)
        if len(ft_weekly) > 0:
            block_a['ft_hours_weekly_mean'] = ft_weekly.mean()
            block_a['ft_hours_weekly_median'] = ft_weekly.median()
            block_a['ft_hours_weekly_min'] = ft_weekly.min()
            block_a['ft_hours_weekly_max'] = ft_weekly.max()
            block_a['ft_hours_weekly_p25'] = ft_weekly.quantile(0.25)
            block_a['ft_hours_weekly_p75'] = ft_weekly.quantile(0.75)
        else:
            for stat in ['mean', 'median', 'min', 'max', 'p25', 'p75']:
                block_a[f'ft_hours_weekly_{stat}'] = np.nan
    else:
        block_a['ft_hours_weekly_nonmissing'] = 0
        for stat in ['mean', 'median', 'min', 'max', 'p25', 'p75']:
            block_a[f'ft_hours_weekly_{stat}'] = np.nan
    
    results.append(block_a)
    
    # Block A2 - Normalized monthly band (relaxed statutory floor + analysis cap) for increase events
    if band_summary is not None and len(band_summary) > 0:
        block_band = {"section": "salary_monthly_band"}
        for key in (
            "n_increase_events",
            "n_conversion_ok",
            "n_band_eligible",
            "n_dropped_missing_salary_date",
            "n_dropped_below_floor",
            "n_dropped_above_cap",
            "share_missing_date_of_conversion_ok",
            "share_below_floor_of_conversion_ok",
            "share_above_cap_of_conversion_ok",
            "share_band_eligible_of_conversion_ok",
        ):
            block_band[key] = band_summary.get(key, np.nan)
        results.append(block_band)
    
    # Block B - Contracts by contract_start_year
    if "contract_start_year" in df.columns:
        for year, year_group in df.groupby("contract_start_year"):
            if pd.isna(year):
                continue
            
            year_row = {
                'section': 'contracts_by_year',
                'contract_start_year': int(year),
                'n_rows_year': len(year_group)
            }
            
            if "cao_number" in year_group.columns:
                year_row['n_cao_year'] = year_group['cao_number'].nunique()
            else:
                year_row['n_cao_year'] = np.nan
            
            results.append(year_row)
    
    # Block C - Salary episodes by salary_start_year
    if len(df_long) > 0 and "salary_start_year" in df_long.columns:
        for year, year_group in df_long.groupby("salary_start_year"):
            if pd.isna(year):
                continue
            
            year_row = {
                'section': 'salary_episodes_by_year',
                'salary_start_year': int(year),
                'n_salary_episodes': len(year_group)
            }
            
            if "cao_number" in year_group.columns:
                year_row['n_cao_year'] = year_group['cao_number'].nunique()
            else:
                year_row['n_cao_year'] = np.nan
            
            if "jobgroup" in year_group.columns:
                year_row['n_jobgroups_year'] = year_group['jobgroup'].nunique()
            else:
                year_row['n_jobgroups_year'] = np.nan
            
            results.append(year_row)
    
    # Block D - CAO-level structure distribution
    if "cao_number" in df.columns:
        cao_stats = []
        for cao_num, cao_group in df.groupby("cao_number"):
            if pd.isna(cao_num):
                continue
            
            cao_stat = {
                'cao_number': cao_num,
                'rows_per_cao': len(cao_group)
            }
            
            for col, label in [("jobgroup", "jobgroups_per_cao"), ("step_label", "steps_per_cao"),
                              ("worker_type", "worker_types_per_cao"), ("age_group", "age_groups_per_cao"),
                              ("education", "educations_per_cao")]:
                if col in cao_group.columns:
                    cao_stat[label] = cao_group[col].nunique()
                else:
                    cao_stat[label] = 0
            
            cao_stats.append(cao_stat)
        
        if len(cao_stats) > 0:
            cao_df = pd.DataFrame(cao_stats)
            for metric in ['rows_per_cao', 'jobgroups_per_cao', 'steps_per_cao',
                          'worker_types_per_cao', 'age_groups_per_cao', 'educations_per_cao']:
                if metric in cao_df.columns:
                    metric_series = cao_df[metric]
                    summary_row = {
                        'section': 'cao_structure_distribution',
                        'metric': metric,
                        'mean': metric_series.mean(),
                        'median': metric_series.median(),
                        'min': metric_series.min(),
                        'max': metric_series.max(),
                        'p25': metric_series.quantile(0.25),
                        'p75': metric_series.quantile(0.75)
                    }
                    results.append(summary_row)
    
    # Block E - FT hours development
    if "contract_start_year" in df.columns and "ft_hours_weekly" in df.columns:
        for year, year_group in df.groupby("contract_start_year"):
            if pd.isna(year):
                continue
            
            ft_weekly_year = year_group["ft_hours_weekly"].dropna()
            if len(ft_weekly_year) > 0:
                year_row = {
                    'section': 'ft_hours_by_year',
                    'contract_start_year': int(year),
                    'n_year': len(ft_weekly_year),
                    'mean_year': ft_weekly_year.mean(),
                    'median_year': ft_weekly_year.median(),
                    'min_year': ft_weekly_year.min(),
                    'max_year': ft_weekly_year.max(),
                    'p25_year': ft_weekly_year.quantile(0.25),
                    'p75_year': ft_weekly_year.quantile(0.75)
                }
                results.append(year_row)
    
    return pd.DataFrame(results)


def create_salary_slots_coverage_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create salary slots coverage sheet.
    
    Args:
        df: Wide format DataFrame
        
    Returns:
        DataFrame with slot coverage statistics
    """
    results = []
    SLOT_RANGE = detect_salary_slot_indices(df.columns.tolist())
    total_rows = len(df)
    
    for k in SLOT_RANGE:
        amount_col = f"salary_{k}_amount"
        if amount_col not in df.columns:
            continue
        
        _coerced = df[amount_col].map(coerce_salary_amount_scalar)
        n_rows_with_amount = int((_coerced.notna() & (_coerced > 0)).sum())
        share_rows_with_amount = n_rows_with_amount / total_rows if total_rows > 0 else 0
        
        start_date_col = f"salary_{k}_start_date"
        if start_date_col in df.columns:
            n_rows_with_start_date = df[start_date_col].notna().sum()
        else:
            n_rows_with_start_date = 0
        
        unit_col = f"salary_{k}_unit"
        if unit_col in df.columns:
            unit_series = df[unit_col].dropna()
            if len(unit_series) > 0:
                most_common_unit = unit_series.mode()[0] if len(unit_series.mode()) > 0 else np.nan
                share_most_common_unit = (unit_series == most_common_unit).sum() / len(unit_series) if len(unit_series) > 0 else 0
            else:
                most_common_unit = np.nan
                share_most_common_unit = 0
        else:
            most_common_unit = np.nan
            share_most_common_unit = 0
        
        results.append({
            'salary_index': k,
            'n_rows_with_amount': n_rows_with_amount,
            'share_rows_with_amount': share_rows_with_amount,
            'n_rows_with_start_date': n_rows_with_start_date,
            'most_common_unit': most_common_unit,
            'share_most_common_unit': share_most_common_unit
        })
    
    return pd.DataFrame(results)


def create_worker_profile_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create worker profile sheet with frequency tables.
    
    Args:
        df: Wide format DataFrame
        
    Returns:
        DataFrame with worker profile statistics
    """
    results = []
    
    # worker_type
    if "worker_type" in df.columns:
        worker_counts = df["worker_type"].value_counts()
        for category, count in worker_counts.items():
            results.append({
                'variable': 'worker_type',
                'category': category,
                'n_rows': count,
                'share_rows': count / len(df) if len(df) > 0 else 0
            })
    
    # permanency
    if "permanency" in df.columns:
        perm_counts = df["permanency"].value_counts()
        for category, count in perm_counts.items():
            results.append({
                'variable': 'permanency',
                'category': category,
                'n_rows': count,
                'share_rows': count / len(df) if len(df) > 0 else 0
            })
    
    # is_entry (boolean)
    if "is_entry" in df.columns:
        bool_series = normalize_boolean(df["is_entry"])
        n_true = (bool_series == True).sum()
        n_false = (bool_series == False).sum()
        results.append({
            'variable': 'is_entry',
            'category': 'True',
            'n_rows': n_true,
            'share_rows': n_true / len(df) if len(df) > 0 else 0
        })
        results.append({
            'variable': 'is_entry',
            'category': 'False',
            'n_rows': n_false,
            'share_rows': n_false / len(df) if len(df) > 0 else 0
        })
    
    # age_group
    if "age_group" in df.columns:
        age_counts = df["age_group"].value_counts()
        for category, count in age_counts.items():
            results.append({
                'variable': 'age_group',
                'category': category,
                'n_rows': count,
                'share_rows': count / len(df) if len(df) > 0 else 0
            })
    
    # education
    if "education" in df.columns:
        edu_counts = df["education"].value_counts()
        for category, count in edu_counts.items():
            results.append({
                'variable': 'education',
                'category': category,
                'n_rows': count,
                'share_rows': count / len(df) if len(df) > 0 else 0
            })
    
    return pd.DataFrame(results)


def create_amounts_by_unit_sheet(df_long: pd.DataFrame) -> pd.DataFrame:
    """
    Create amounts by unit sheet.
    
    Args:
        df_long: Long format DataFrame
        
    Returns:
        DataFrame with amount statistics by unit
    """
    results = []
    
    if len(df_long) == 0:
        return pd.DataFrame()
    
    # Filter to non-NaN salary_amount and salary_unit
    df_filtered = df_long[
        df_long["salary_amount"].notna() & df_long["salary_unit"].notna()
    ].copy()
    
    if len(df_filtered) == 0:
        return pd.DataFrame()
    
    # Coerce salary_amount to numeric
    df_filtered["salary_amount"] = pd.to_numeric(df_filtered["salary_amount"], errors='coerce')
    df_filtered = df_filtered[df_filtered["salary_amount"].notna()]
    
    if len(df_filtered) == 0:
        return pd.DataFrame()
    
    # Block A - By (salary_unit, salary_index)
    if "salary_index" in df_filtered.columns:
        for (unit, idx), group in df_filtered.groupby(["salary_unit", "salary_index"]):
            amounts = group["salary_amount"]
            results.append({
                'block': 'by_unit_and_index',
                'salary_unit': unit,
                'salary_index': idx,
                'n': len(amounts),
                'mean': amounts.mean(),
                'median': amounts.median(),
                'p25': amounts.quantile(0.25),
                'p75': amounts.quantile(0.75),
                'min': amounts.min(),
                'max': amounts.max()
            })
    
    # Block B - By salary_unit (aggregated)
    for unit, group in df_filtered.groupby("salary_unit"):
        amounts = group["salary_amount"]
        results.append({
            'block': 'by_unit',
            'salary_unit': unit,
            'salary_index': np.nan,
            'n': len(amounts),
            'mean': amounts.mean(),
            'median': amounts.median(),
            'p25': amounts.quantile(0.25),
            'p75': amounts.quantile(0.75),
            'min': amounts.min(),
            'max': amounts.max()
        })
    
    return pd.DataFrame(results)


def create_increase_percent_sheet(df_long: pd.DataFrame) -> pd.DataFrame:
    """
    Create increase percent sheet.
    
    Args:
        df_long: Long format DataFrame
        
    Returns:
        DataFrame with increase percent statistics
    """
    results = []
    
    if len(df_long) == 0 or "salary_increase_percent" not in df_long.columns:
        return pd.DataFrame(columns=['block', 'statistic', 'value'])
    
    # Coerce to numeric
    s_inc = pd.to_numeric(df_long["salary_increase_percent"], errors='coerce')
    s_inc_clean = s_inc.dropna()
    
    if len(s_inc_clean) == 0:
        return pd.DataFrame(columns=['block', 'statistic', 'value'])
    
    # Block A - Overall distribution
    results.append({
        'block': 'overall',
        'statistic': 'n',
        'value': len(s_inc_clean)
    })
    results.append({
        'block': 'overall',
        'statistic': 'mean',
        'value': s_inc_clean.mean()
    })
    results.append({
        'block': 'overall',
        'statistic': 'median',
        'value': s_inc_clean.median()
    })
    results.append({
        'block': 'overall',
        'statistic': 'p25',
        'value': s_inc_clean.quantile(0.25)
    })
    results.append({
        'block': 'overall',
        'statistic': 'p75',
        'value': s_inc_clean.quantile(0.75)
    })
    results.append({
        'block': 'overall',
        'statistic': 'min',
        'value': s_inc_clean.min()
    })
    results.append({
        'block': 'overall',
        'statistic': 'max',
        'value': s_inc_clean.max()
    })
    
    # Block B - By salary_start_year
    if "salary_start_year" in df_long.columns:
        df_long_inc = df_long.copy()
        df_long_inc["salary_increase_percent_numeric"] = s_inc
        
        for year, year_group in df_long_inc.groupby("salary_start_year"):
            if pd.isna(year):
                continue
            
            year_inc = year_group["salary_increase_percent_numeric"].dropna()
            if len(year_inc) >= 3:
                results.append({
                    'block': 'by_year',
                    'statistic': 'year',
                    'value': int(year)
                })
                results.append({
                    'block': 'by_year',
                    'statistic': 'n_year',
                    'value': len(year_inc)
                })
                results.append({
                    'block': 'by_year',
                    'statistic': 'mean_year',
                    'value': year_inc.mean()
                })
                results.append({
                    'block': 'by_year',
                    'statistic': 'median_year',
                    'value': year_inc.median()
                })
    
    return pd.DataFrame(results)


def create_diff_only_increase_sheet(events_df: pd.DataFrame) -> pd.DataFrame:
    """
    Create yearly summary for increase_diff_only.

    Args:
        events_df: Event-level salary DataFrame from derivation module

    Returns:
        DataFrame with summary rows and yearly aggregates
    """
    if len(events_df) == 0 or "increase_diff_only" not in events_df.columns:
        return pd.DataFrame()
    out: List[Dict[str, Any]] = []
    valid = events_df[events_df["increase_diff_only"].notna()].copy()
    out.append(
        {
            "section": "summary",
            "metric": "n_events_total",
            "value": len(events_df),
        }
    )
    out.append({"section": "summary", "metric": "n_events_diff_only", "value": len(valid)})
    out.append({"section": "summary", "metric": "share_with_diff_only", "value": len(valid) / len(events_df) if len(events_df) else np.nan})
    if len(valid) > 0:
        out.append({"section": "summary", "metric": "year_min", "value": int(valid["salary_start_year"].min())})
        out.append({"section": "summary", "metric": "year_max", "value": int(valid["salary_start_year"].max())})
        yearly = valid.groupby("salary_start_year")["increase_diff_only"].agg(["count", "mean", "median"]).reset_index()
        for _, row in yearly.iterrows():
            out.append(
                {
                    "section": "yearly_diff_only",
                    "salary_start_year": int(row["salary_start_year"]),
                    "n": int(row["count"]),
                    "mean_increase_percent": row["mean"],
                    "median_increase_percent": row["median"],
                    "note": "Amounts normalized to monthly before differencing.",
                }
            )
    return pd.DataFrame(out)


def create_merge_vs_csv_sheet(comparison_df: pd.DataFrame) -> pd.DataFrame:
    """
    Create comparison sheet for CSV vs derived diff increases.

    Args:
        comparison_df: Rows where both CSV and diff increases exist

    Returns:
        DataFrame with global metrics, per-year appendix and top differences
    """
    if len(comparison_df) == 0:
        return pd.DataFrame()
    rows: List[Dict[str, Any]] = []
    rows.append({"section": "summary", "metric": "n_comparable_rows", "value": len(comparison_df)})
    rows.append({"section": "summary", "metric": "agreement_share_abs_diff_le_0_1pp", "value": comparison_df["within_0_1pp"].mean()})
    rows.append({"section": "summary", "metric": "sign_disagreement_share", "value": comparison_df["sign_disagreement"].mean()})

    by_year = comparison_df.groupby("salary_start_year").agg(
        n=("within_0_1pp", "count"),
        agreement_share=("within_0_1pp", "mean"),
        mean_abs_diff=("abs_diff", "mean"),
    ).reset_index()
    for _, row in by_year.iterrows():
        rows.append(
            {
                "section": "per_year_agreement",
                "salary_start_year": int(row["salary_start_year"]),
                "n": int(row["n"]),
                "agreement_share_abs_diff_le_0_1pp": row["agreement_share"],
                "mean_abs_diff": row["mean_abs_diff"],
            }
        )

    top25 = comparison_df.sort_values("abs_diff", ascending=False).head(25)
    for _, row in top25.iterrows():
        rows.append(
            {
                "section": "top_25_abs_diff",
                "cao_number": row.get("cao_number"),
                "file_name": row.get("file_name"),
                "row_id": row.get("row_id"),
                "salary_index": row.get("salary_index"),
                "salary_start_date": row.get("salary_start_date"),
                "increase_csv_only": row.get("increase_csv_only"),
                "increase_diff_only": row.get("increase_diff_only"),
                "abs_diff": row.get("abs_diff"),
                "sign_disagreement": row.get("sign_disagreement"),
            }
        )
    return pd.DataFrame(rows)


def create_level_trends_by_year_sheet(df_long: pd.DataFrame) -> pd.DataFrame:
    """
    Create level trends by year sheet.
    
    Args:
        df_long: Long format DataFrame
        
    Returns:
        DataFrame with level trends statistics
    """
    results = []
    
    if len(df_long) == 0:
        return pd.DataFrame()
    
    # Identify primary_unit
    if "salary_unit" in df_long.columns:
        unit_series = df_long["salary_unit"].dropna()
        if len(unit_series) > 0:
            primary_unit = unit_series.mode()[0] if len(unit_series.mode()) > 0 else None
        else:
            primary_unit = None
    else:
        primary_unit = None
    
    if primary_unit is None:
        return pd.DataFrame()
    
    # Filter to primary_unit rows with non-NaN salary_amount and salary_start_year
    df_filtered = df_long[
        (df_long["salary_unit"] == primary_unit) &
        df_long["salary_amount"].notna() &
        df_long["salary_start_year"].notna()
    ].copy()
    
    if len(df_filtered) == 0:
        return pd.DataFrame()
    
    # Coerce salary_amount to numeric
    df_filtered["salary_amount"] = pd.to_numeric(df_filtered["salary_amount"], errors='coerce')
    df_filtered = df_filtered[df_filtered["salary_amount"].notna()]
    
    if len(df_filtered) == 0:
        return pd.DataFrame()
    
    # Block A - All episodes in primary_unit
    for year, year_group in df_filtered.groupby("salary_start_year"):
        if pd.isna(year):
            continue
        
        amounts = year_group["salary_amount"]
        results.append({
            'block': 'all_episodes',
            'salary_start_year': int(year),
            'primary_unit': primary_unit,
            'n_episodes': len(amounts),
            'mean_amount': amounts.mean(),
            'median_amount': amounts.median(),
            'p25_amount': amounts.quantile(0.25),
            'p75_amount': amounts.quantile(0.75)
        })
    
    # Block B - Entry scales only
    if "is_entry" in df_filtered.columns:
        df_entry = df_filtered[normalize_boolean(df_filtered["is_entry"]) == True].copy()
        
        if len(df_entry) > 0:
            for year, year_group in df_entry.groupby("salary_start_year"):
                if pd.isna(year):
                    continue
                
                amounts = year_group["salary_amount"]
                results.append({
                    'block': 'entry_scales',
                    'salary_start_year': int(year),
                    'primary_unit': primary_unit,
                    'n_entry_episodes': len(amounts),
                    'mean_entry_amount': amounts.mean(),
                    'median_entry_amount': amounts.median()
                })
    
    return pd.DataFrame(results)


def create_row_slot_counts_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create row slot counts sheet.
    
    Args:
        df: Wide format DataFrame with n_salary_points_per_row
        
    Returns:
        DataFrame with row slot count statistics
    """
    results = []
    
    if "n_salary_points_per_row" not in df.columns:
        return pd.DataFrame()
    
    total_rows = len(df)
    
    # Block A - Overall distribution: counts 0..10 plus 11+ (>=11 salary points)
    for n in range(0, 11):
        n_rows_n = (df["n_salary_points_per_row"] == n).sum()
        share_rows_n = n_rows_n / total_rows if total_rows > 0 else 0
        results.append({
            'block': 'overall',
            'n_salary_points_per_row': n,
            'n_rows': int(n_rows_n),
            'share_rows': share_rows_n
        })
    n_rows_ge_11 = (df["n_salary_points_per_row"] >= 11).sum()
    share_ge_11 = n_rows_ge_11 / total_rows if total_rows > 0 else 0
    results.append({
        'block': 'overall',
        'n_salary_points_per_row': '11+',
        'n_rows': int(n_rows_ge_11),
        'share_rows': share_ge_11
    })
    
    # Block B - By contract_start_year
    if "contract_start_year" in df.columns:
        for year, year_group in df.groupby("contract_start_year"):
            if pd.isna(year):
                continue
            
            year_total = len(year_group)
            for n in range(0, 11):
                n_rows_year_n = ((year_group["n_salary_points_per_row"] == n).sum())
                share_rows_year_n = n_rows_year_n / year_total if year_total > 0 else 0
                results.append({
                    'block': 'by_year',
                    'contract_start_year': int(year),
                    'n_salary_points_per_row': n,
                    'n_rows': int(n_rows_year_n),
                    'share_rows': share_rows_year_n
                })
            n_y_ge_11 = (year_group["n_salary_points_per_row"] >= 11).sum()
            share_y_ge_11 = n_y_ge_11 / year_total if year_total > 0 else 0
            results.append({
                'block': 'by_year',
                'contract_start_year': int(year),
                'n_salary_points_per_row': '11+',
                'n_rows': int(n_y_ge_11),
                'share_rows': share_y_ge_11
            })
    
    return pd.DataFrame(results)


def create_no_salary_analysis_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create analysis sheet for files with no salary data.
    
    Analyzes files that have no salary information (n_salary_points_per_row == 0).
    
    Args:
        df: Wide format DataFrame
        
    Returns:
        DataFrame with no salary analysis statistics
    """
    results = []
    
    if "n_salary_points_per_row" not in df.columns:
        return pd.DataFrame()
    
    # Local mask only — do not mutate the wide frame (avoids further fragmentation).
    has_no_salary = df["n_salary_points_per_row"] == 0
    
    # Block A - Overall summary
    total_files = df[['cao_number', 'file_name']].drop_duplicates().shape[0] if 'cao_number' in df.columns and 'file_name' in df.columns else len(df)
    if 'cao_number' in df.columns and 'file_name' in df.columns:
        files_with_no_salary = df.loc[has_no_salary, ['cao_number', 'file_name']].drop_duplicates().shape[0]
    else:
        files_with_no_salary = int(has_no_salary.sum())
    files_with_salary = total_files - files_with_no_salary
    
    results.append({
        'block': 'overall_summary',
        'metric': 'total_unique_files',
        'value': total_files,
        'share': 1.0
    })
    results.append({
        'block': 'overall_summary',
        'metric': 'files_with_salary_data',
        'value': files_with_salary,
        'share': files_with_salary / total_files if total_files > 0 else 0
    })
    results.append({
        'block': 'overall_summary',
        'metric': 'files_with_no_salary_data',
        'value': files_with_no_salary,
        'share': files_with_no_salary / total_files if total_files > 0 else 0
    })
    
    # Block B - By CAO number
    if 'cao_number' in df.columns and 'file_name' in df.columns:
        # Count files with/without salary per CAO (narrow frame; avoid adding column to wide df)
        cao_file_summary = (
            df[['cao_number', 'file_name']]
            .assign(has_no_salary=has_no_salary)
            .groupby(['cao_number', 'file_name'], sort=False)
            .agg({'has_no_salary': 'any'})
            .reset_index()
        )
        
        # Count files per CAO
        cao_summary = cao_file_summary.groupby('cao_number').agg({
            'has_no_salary': ['sum', 'count']  # sum = count of no-salary files, count = total files
        }).reset_index()
        cao_summary.columns = ['cao_number', 'n_files_no_salary', 'n_files_total']
        cao_summary['n_files_with_salary'] = cao_summary['n_files_total'] - cao_summary['n_files_no_salary']
        cao_summary['has_only_no_salary'] = cao_summary['n_files_with_salary'] == 0
        cao_summary['has_any_salary'] = cao_summary['n_files_with_salary'] > 0
        
        # CAOs with ONLY no-salary files (no salary data at all)
        caos_only_no_salary = cao_summary[cao_summary['has_only_no_salary']]
        # CAOs with at least some salary files
        caos_with_salary = cao_summary[cao_summary['has_any_salary']]
        # CAOs with at least one no-salary file (but may also have salary files)
        caos_with_some_no_salary = cao_summary[cao_summary['n_files_no_salary'] > 0]
        
        results.append({
            'block': 'by_cao',
            'metric': 'caos_with_only_no_salary_files',
            'value': len(caos_only_no_salary),
            'share': len(caos_only_no_salary) / len(cao_summary) if len(cao_summary) > 0 else 0
        })
        results.append({
            'block': 'by_cao',
            'metric': 'caos_with_at_least_some_salary_files',
            'value': len(caos_with_salary),
            'share': len(caos_with_salary) / len(cao_summary) if len(cao_summary) > 0 else 0
        })
        results.append({
            'block': 'by_cao',
            'metric': 'caos_with_at_least_one_no_salary_file',
            'value': len(caos_with_some_no_salary),
            'share': len(caos_with_some_no_salary) / len(cao_summary) if len(cao_summary) > 0 else 0
        })
        results.append({
            'block': 'by_cao',
            'metric': 'total_files_in_caos_with_only_no_salary',
            'value': caos_only_no_salary['n_files_total'].sum() if len(caos_only_no_salary) > 0 else 0,
            'share': caos_only_no_salary['n_files_total'].sum() / cao_summary['n_files_total'].sum() if cao_summary['n_files_total'].sum() > 0 else 0
        })
    
    # Block C - By contract_start_year
    if 'contract_start_year' in df.columns:
        year_helper = df[['contract_start_year', 'cao_number', 'file_name']].assign(
            has_no_salary=has_no_salary
        )
        year_summary = year_helper.groupby('contract_start_year', sort=False).agg({
            'has_no_salary': 'sum',
            'cao_number': 'nunique',
            'file_name': 'nunique',
        }).reset_index()
        year_summary.columns = ['contract_start_year', 'n_rows_no_salary', 'n_cao', 'n_files']
        # Get total rows per year
        year_counts = df.groupby('contract_start_year').size().reset_index(name='n_rows_total')
        year_summary = year_summary.merge(year_counts, on='contract_start_year', how='left')
        year_summary['share_rows_no_salary'] = year_summary['n_rows_no_salary'] / year_summary['n_rows_total']
        
        for _, row in year_summary.iterrows():
            if pd.notna(row['contract_start_year']):
                results.append({
                    'block': 'by_contract_year',
                    'contract_start_year': int(row['contract_start_year']),
                    'n_files': int(row['n_files']),
                    'n_rows_no_salary': int(row['n_rows_no_salary']),
                    'n_rows_total': int(row['n_rows_total']),
                    'share_rows_no_salary': row['share_rows_no_salary'],
                    'n_cao': int(row['n_cao'])
                })
    
    # Block D - List of CAOs with no salary files (if any)
    if 'cao_number' in df.columns and 'file_name' in df.columns:
        no_salary_files = df.loc[has_no_salary, ['cao_number', 'file_name']].drop_duplicates()
        if len(no_salary_files) > 0:
            # Add metadata if available
            for col in ['id', 'TTW', 'ingangsdatum', 'expiratiedatum', 'sector', 'sbi_code']:
                if col in df.columns:
                    # Get first value for each (cao_number, file_name) combination
                    metadata = df.loc[has_no_salary].groupby(['cao_number', 'file_name'])[col].first().reset_index()
                    no_salary_files = no_salary_files.merge(metadata, on=['cao_number', 'file_name'], how='left')
            
            # Add to results as a separate block
            for _, row in no_salary_files.iterrows():
                result_row = {
                    'block': 'no_salary_files_list',
                    'cao_number': row['cao_number'],
                    'file_name': row['file_name']
                }
                # Add available metadata
                for col in ['id', 'TTW', 'ingangsdatum', 'expiratiedatum', 'sector', 'sbi_code']:
                    if col in row:
                        result_row[col] = row[col]
                results.append(result_row)
    
    return pd.DataFrame(results)


def build_salary_descriptives_exclusion_report(df: pd.DataFrame) -> pd.DataFrame:
    """
    Build a combined summary and file list for wide rows and files excluded from the salary analytic sample.

    The analytic sample keeps rows with n_salary_points_per_row > 0. Detail rows list (cao_number, file_name)
    pairs where every wide row in that file has zero salary points (fully no-salary contract files).

    Args:
        df: Full wide frame after n_salary_points_per_row was attached.

    Returns:
        DataFrame with record_type 'summary' (one row) and 'excluded_file' (one row per fully no-salary file).
    """
    rows: List[Dict[str, Any]] = []
    if "n_salary_points_per_row" not in df.columns:
        return pd.DataFrame(rows)

    n_total = len(df)
    has_salary = df["n_salary_points_per_row"].to_numpy() > 0
    n_analytic = int(has_salary.sum())
    n_excluded = n_total - n_analytic

    n_unique_files_total = np.nan
    n_unique_files_fully_no_salary = np.nan
    n_cao_total = np.nan
    n_cao_with_at_least_one_fully_no_salary_file = np.nan
    n_cao_all_files_fully_no_salary = np.nan
    fully_no_salary_keys: List[Any] = []

    if "cao_number" in df.columns and "file_name" in df.columns:
        n_unique_files_total = int(df[["cao_number", "file_name"]].drop_duplicates().shape[0])
        n_cao_total = int(df["cao_number"].nunique())
        file_max = df.groupby(["cao_number", "file_name"], sort=False)["n_salary_points_per_row"].max()
        mask_fully = file_max == 0
        fully_no_salary_keys = list(file_max.loc[mask_fully].index)
        n_unique_files_fully_no_salary = len(fully_no_salary_keys)
        if fully_no_salary_keys:
            cao_vals = [t[0] for t in fully_no_salary_keys]
            n_cao_with_at_least_one_fully_no_salary_file = len(set(cao_vals))
        else:
            n_cao_with_at_least_one_fully_no_salary_file = 0

        cf = (
            df.groupby(["cao_number", "file_name"], sort=False)["n_salary_points_per_row"]
            .max()
            .reset_index()
        )
        cf["fully_no_salary_file"] = cf["n_salary_points_per_row"] == 0
        g = cf.groupby("cao_number", sort=False).agg(
            n_files=("file_name", "count"), n_fully_no_salary=("fully_no_salary_file", "sum")
        )
        n_cao_all_files_fully_no_salary = int((g["n_fully_no_salary"] == g["n_files"]).sum())

    rows.append(
        {
            "record_type": "summary",
            "n_wide_rows_total": n_total,
            "n_wide_rows_analytic": n_analytic,
            "n_wide_rows_excluded": n_excluded,
            "n_unique_files_total": n_unique_files_total,
            "n_unique_files_fully_no_salary": n_unique_files_fully_no_salary,
            "n_cao_total": n_cao_total,
            "n_cao_with_at_least_one_fully_no_salary_file": n_cao_with_at_least_one_fully_no_salary_file,
            "n_cao_all_files_fully_no_salary": n_cao_all_files_fully_no_salary,
        }
    )

    meta_cols = [c for c in ("id", "TTW", "ingangsdatum", "expiratiedatum", "sector", "sbi_code") if c in df.columns]
    for key in fully_no_salary_keys:
        cao, fn = key[0], key[1]
        drow: Dict[str, Any] = {
            "record_type": "excluded_file",
            "cao_number": cao,
            "file_name": fn,
        }
        sub = df.loc[(df["cao_number"] == cao) & (df["file_name"] == fn)]
        if len(sub) > 0 and meta_cols:
            first = sub.iloc[0]
            for c in meta_cols:
                drow[c] = first.get(c, np.nan)
        rows.append(drow)

    return pd.DataFrame(rows)


# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    """Main entry point for descriptives script."""
    # Wide read_csv frames stay multi-block; we batch new column writes, but pandas may still warn.
    # Avoid a full-frame copy (RAM); suppress only this specific fragmentation PerformanceWarning.
    warnings.filterwarnings(
        "ignore",
        message="DataFrame is highly fragmented",
        category=pd.errors.PerformanceWarning,
    )
    print("="*80)
    print("CAO Salary Descriptives Script")
    print("="*80)
    
    # Load data
    print(f"\nLoading data from: {INPUT_CSV}")
    try:
        df = pd.read_csv(INPUT_CSV, sep=';', encoding='utf-8', low_memory=False)
        print(f"  Loaded {len(df)} rows and {len(df.columns)} columns")
        log_memory("raw_wide", df)
    except Exception as e:
        print(f"  ERROR: Could not load input file: {e}")
        return
    
    if len(df) == 0:
        print("  ERROR: Input file is empty")
        return
    
    # Parse date columns (CAO metadata dates are DD/MM/YYYY; use robust parser)
    date_cols = ["ingangsdatum", "expiratiedatum", "datum_kennisgeving"]
    date_cols_present = [c for c in date_cols if c in df.columns]
    for col in date_cols:
        if col not in df.columns:
            print(f"  Warning: {col} column not found")
    if date_cols_present:
        df[date_cols_present] = df[date_cols_present].apply(
            lambda s: parse_cao_date_series(s, dayfirst=True)
        )
        for col in date_cols_present:
            print(f"  Parsed {col} as datetime")
    
    # Parse salary date columns
    SLOT_RANGE = detect_salary_slot_indices(df.columns.tolist())
    if SLOT_RANGE:
        print(
            f"  Detected salary slots: count={len(SLOT_RANGE)}, "
            f"index_min={SLOT_RANGE[0]}, index_max={SLOT_RANGE[-1]}"
        )
    else:
        print("  Detected salary slots: none (no salary_<k>_amount columns matched)")
    slot_date_cols = [
        c
        for k in SLOT_RANGE
        for c in (f"salary_{k}_start_date", f"salary_{k}_end_date")
        if c in df.columns
    ]
    if slot_date_cols:
        df[slot_date_cols] = df[slot_date_cols].apply(
            lambda s: pd.to_datetime(s, errors="coerce")
        )
    
    # Create contract_start_year and ft_hours_weekly in one block assignment (less fragmentation)
    derived_cols: Dict[str, Any] = {}
    if "ingangsdatum" in df.columns:
        derived_cols["contract_start_year"] = df["ingangsdatum"].dt.year
    else:
        print("  Warning: ingangsdatum not found, cannot create contract_start_year")
        derived_cols["contract_start_year"] = np.nan
    if "ft_hours" in df.columns:
        ft_hours_numeric = pd.to_numeric(df["ft_hours"], errors="coerce")
        v = ft_hours_numeric.to_numpy(dtype=float, copy=False)
        with np.errstate(invalid="ignore"):
            derived_cols["ft_hours_weekly"] = np.where(
                np.isnan(v), np.nan, np.where(v <= 200, v, v / 52.0)
            )
        print("  Created ft_hours_weekly from ft_hours")
    else:
        print("  Warning: ft_hours column not found")
        derived_cols["ft_hours_weekly"] = np.nan
    df["contract_start_year"] = derived_cols["contract_start_year"]
    df["ft_hours_weekly"] = derived_cols["ft_hours_weekly"]

    # n_salary_points_per_row before long/increase (strictly positive coerced amounts only)
    print("\nComputing n_salary_points_per_row...")
    n_rows = len(df)
    n_points = np.zeros(n_rows, dtype=np.int32)
    for k in SLOT_RANGE:
        amount_col = f"salary_{k}_amount"
        if amount_col in df.columns:
            _am = df[amount_col].map(coerce_salary_amount_scalar)
            n_points += (_am.notna() & (_am > 0)).astype(np.int32).to_numpy(copy=False)
    df["n_salary_points_per_row"] = n_points

    output_path = Path(OUTPUT_EXCEL)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    diagnostics_dir = Path("outputs/analysis")
    diagnostics_dir.mkdir(parents=True, exist_ok=True)

    exclusion_report_df = build_salary_descriptives_exclusion_report(df)
    exclusion_csv_path = diagnostics_dir / Path(OUTPUT_EXCLUSION_CSV).name
    exclusion_report_df.to_csv(exclusion_csv_path, index=False, sep=";", decimal=",")
    print(f"\nWrote {exclusion_csv_path} ({len(exclusion_report_df)} rows)")

    has_salary_mask = df["n_salary_points_per_row"].to_numpy() > 0
    source_row_ids = np.flatnonzero(has_salary_mask)
    df_analytic = df.loc[has_salary_mask].reset_index(drop=True)
    if len(df_analytic) > 0:
        df_analytic["_wide_source_row_id"] = source_row_ids
    print(
        f"  Analytic wide sample: {len(df_analytic)} rows "
        f"(excluded {int(n_rows - len(df_analytic))} wide rows with no positive salary amount)"
    )

    # Build long format DataFrame
    print("\nBuilding long format DataFrame...")
    try:
        df_long = build_long_salary_df(df_analytic)
        print(f"  Long format: {len(df_long)} rows")
        log_memory("long_salary", df_long)
    except Exception as e:
        print(f"  Warning: Error building long format: {e}")
        df_long = pd.DataFrame()

    print("\nDeriving salary increase series (diff/csv/merged)...")
    try:
        increase_payload = derive_salary_increase_series(df_analytic)
        increase_events = increase_payload["events"]
        comparison_df = increase_payload["comparison"]
        conversion_diag = increase_payload["conversion_diagnostics"]
        band_summary = increase_payload.get("band_summary")
        if band_summary is None:
            band_summary = compute_band_summary_stats(increase_events)
        print(f"  Derived events: {len(increase_events)} | comparable rows: {len(comparison_df)}")
        if band_summary:
            print(
                f"  Monthly band: eligible={band_summary.get('n_band_eligible', 0)} "
                f"| below_floor={band_summary.get('n_dropped_below_floor', 0)} "
                f"| above_cap={band_summary.get('n_dropped_above_cap', 0)} "
                f"| missing_date={band_summary.get('n_dropped_missing_salary_date', 0)} "
                f"(conversion_ok={band_summary.get('n_conversion_ok', 0)})"
            )
        if len(increase_events) > 0:
            log_memory("increase_events", increase_events)
    except Exception as e:
        print(f"  Warning: Error deriving salary increase series: {e}")
        increase_events = pd.DataFrame()
        comparison_df = pd.DataFrame()
        conversion_diag = pd.DataFrame()
        band_summary = compute_band_summary_stats(pd.DataFrame())

    df_descriptives_wide = df_analytic.drop(columns=["_wide_source_row_id"], errors="ignore")
    comparison_csv_path = diagnostics_dir / "salary_increase_csv_vs_diff_comparison.csv"
    comparison_df.to_csv(comparison_csv_path, index=False, sep=";", decimal=",")
    print(f"\nWrote {comparison_csv_path} ({len(comparison_df)} rows)")
    band_summary_path = diagnostics_dir / "salary_monthly_band_summary.csv"
    pd.DataFrame([band_summary]).to_csv(band_summary_path, index=False, sep=";", decimal=",")
    print(f"Wrote {band_summary_path}")
    conversion_diag.to_csv(
        diagnostics_dir / "salary_increase_conversion_diagnostics.csv", index=False, sep=";", decimal=","
    )
    print(f"Wrote {diagnostics_dir / 'salary_increase_conversion_diagnostics.csv'} ({len(conversion_diag)} rows)")
    increase_events.to_csv(
        diagnostics_dir / "salary_increase_events_derived.csv", index=False, sep=";", decimal=","
    )
    print(f"Wrote {diagnostics_dir / 'salary_increase_events_derived.csv'} ({len(increase_events)} rows)")
    
    # Generate all sheets (00–07: analytic wide sample only; 08: full extract)
    print("\nGenerating descriptive statistics sheets...")
    
    sheets = {}
    
    try:
        print("  Creating sheet: 00_excluded_no_salary_summary")
        sheets["00_excluded_no_salary_summary"] = exclusion_report_df
    except Exception as e:
        print(f"  Warning: Error creating excluded_no_salary_summary sheet: {e}")
        sheets["00_excluded_no_salary_summary"] = pd.DataFrame()
    
    try:
        print("  Creating sheet: 01_variable_health")
        sheets["01_variable_health"] = create_variable_health_sheet(df_descriptives_wide)
    except Exception as e:
        print(f"  Warning: Error creating variable_health sheet: {e}")
        sheets["01_variable_health"] = pd.DataFrame()
    
    try:
        print("  Creating sheet: 02_sample_overview")
        sheets["02_sample_overview"] = create_sample_overview_sheet(
            df_descriptives_wide, df_long, band_summary=band_summary
        )
    except Exception as e:
        print(f"  Warning: Error creating sample_overview sheet: {e}")
        sheets["02_sample_overview"] = pd.DataFrame()
    
    try:
        print("  Creating sheet: 03_salary_slots_coverage")
        sheets["03_salary_slots_coverage"] = create_salary_slots_coverage_sheet(df_descriptives_wide)
    except Exception as e:
        print(f"  Warning: Error creating salary_slots_coverage sheet: {e}")
        sheets["03_salary_slots_coverage"] = pd.DataFrame()
    
    try:
        print("  Creating sheet: 04_worker_profile")
        sheets["04_worker_profile"] = create_worker_profile_sheet(df_descriptives_wide)
    except Exception as e:
        print(f"  Warning: Error creating worker_profile sheet: {e}")
        sheets["04_worker_profile"] = pd.DataFrame()
    
    try:
        print("  Creating sheet: 05_amounts_by_unit")
        sheets["05_amounts_by_unit"] = create_amounts_by_unit_sheet(df_long)
    except Exception as e:
        print(f"  Warning: Error creating amounts_by_unit sheet: {e}")
        sheets["05_amounts_by_unit"] = pd.DataFrame()
    
    try:
        print("  Creating sheet: 06_increase_percent")
        sheets["06_increase_percent"] = create_increase_percent_sheet(df_long)
    except Exception as e:
        print(f"  Warning: Error creating increase_percent sheet: {e}")
        sheets["06_increase_percent"] = pd.DataFrame()
    
    try:
        print("  Creating sheet: 07_level_trends_by_year")
        sheets["07_level_trends_by_year"] = create_level_trends_by_year_sheet(df_long)
    except Exception as e:
        print(f"  Warning: Error creating level_trends_by_year sheet: {e}")
        sheets["07_level_trends_by_year"] = pd.DataFrame()
    
    try:
        print("  Creating sheet: 08_row_slot_counts")
        sheets["08_row_slot_counts"] = create_row_slot_counts_sheet(df_descriptives_wide)
    except Exception as e:
        print(f"  Warning: Error creating row_slot_counts sheet: {e}")
        sheets["08_row_slot_counts"] = pd.DataFrame()
    
    try:
        print("  Creating sheet: 09_no_salary_analysis")
        sheets["09_no_salary_analysis"] = create_no_salary_analysis_sheet(df)
    except Exception as e:
        print(f"  Warning: Error creating no_salary_analysis sheet: {e}")
        sheets["09_no_salary_analysis"] = pd.DataFrame()

    try:
        print("  Creating sheet: 10_increase_diff_only")
        sheets["10_increase_diff_only"] = create_diff_only_increase_sheet(increase_events)
    except Exception as e:
        print(f"  Warning: Error creating increase_diff_only sheet: {e}")
        sheets["10_increase_diff_only"] = pd.DataFrame()

    try:
        print("  Creating sheet: 11_increase_merge_vs_csv")
        sheets["11_increase_merge_vs_csv"] = create_merge_vs_csv_sheet(comparison_df)
    except Exception as e:
        print(f"  Warning: Error creating increase_merge_vs_csv sheet: {e}")
        sheets["11_increase_merge_vs_csv"] = pd.DataFrame()
    
    # Write to Excel with explanatory notes
    print(f"\nWriting to Excel: {OUTPUT_EXCEL}")
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font
        
        # First write all dataframes
        with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl') as writer:
            for sheet_name, sheet_df in sheets.items():
                if len(sheet_df) > 0:
                    # Replace NaN with None for Excel compatibility
                    sheet_df_clean = sheet_df.where(pd.notna(sheet_df), None)
                    # Convert any remaining problematic types
                    for col in sheet_df_clean.columns:
                        if sheet_df_clean[col].dtype == 'object':
                            # Replace any remaining NaN/NaT with empty string
                            sheet_df_clean[col] = sheet_df_clean[col].fillna('')
                    sheet_df_clean.to_excel(writer, sheet_name=sheet_name, index=False)
                    print(f"  Wrote sheet '{sheet_name}' with {len(sheet_df_clean)} rows")
                else:
                    # Write empty sheet with headers if possible
                    sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    print(f"  Wrote empty sheet '{sheet_name}'")
        
        # Now add explanatory notes using openpyxl
        wb = load_workbook(OUTPUT_EXCEL)
        
        # Add notes to specific sheets
        notes = {
            "00_excluded_no_salary_summary": [
                "NOTES:",
                "record_type=summary: one row with counts for the full extract vs the analytic wide sample.",
                "record_type=excluded_file: one row per (cao_number, file_name) where every wide row in that file has",
                "  n_salary_points_per_row == 0 (no positive coerced salary_k_amount on any job row).",
                "Mirrors outputs/analysis/salary_descriptives_excluded_no_salary.csv (semicolon-separated).",
                "Sheets 01–08 use only wide rows with at least one positive salary amount; sheet 09 uses the full extract.",
            ],
            "01_variable_health": [
                "NOTES:",
                "This sheet provides a comprehensive overview of variables on the analytic wide sample (salary-bearing rows only).",
                "",
                "inferred_type: Automatically detected type (boolean, numeric, date, categorical, text, id/other)",
                "n_nonmissing: Number of non-missing values",
                "share_nonmissing: Proportion of rows with non-missing values",
                "n_distinct: Number of distinct values",
                "",
                "Type-specific statistics:",
                "  - Boolean: n_true, share_true, n_false, share_false, n_missing",
                "  - Numeric: mean, median, min, max, p25, p75, share_zero",
                "  - Categorical: top_category, top_category_share",
                "  - Date: min_date, max_date",
                "  - Text: avg_char_length"
            ],
            "02_sample_overview": [
                "NOTES:",
                "This sheet contains multiple sections (analytic wide sample only — rows with ≥1 positive salary amount):",
                "",
                "1. Overall dataset structure (section='overall_structure'):",
                "   - n_rows: Wide rows in the analytic sample (one row per jobgroup/step/worker combination with salary)",
                "   - n_unique_files: Unique (cao_number + file_name) among analytic rows (files can appear without all job rows)",
                "   - n_cao: Number of unique CAOs",
                "   - n_jobgroups, n_steps, n_worker_types, n_age_groups, n_education_categories: Counts of distinct values",
                "   - ft_hours_weekly: Statistics for full-time weekly hours (converted from annual if >200)",
                "",
                "2. Contracts by contract start year (section='contracts_by_year'):",
                "   - n_rows_year: Number of salary rows with contracts starting in that year",
                "   - n_cao_year: Number of unique CAOs with contracts starting in that year",
                "",
                "3. Salary episodes by salary start year (section='salary_episodes_by_year'):",
                "   - n_salary_episodes: Number of salary timeline points (episodes) starting in that year",
                "   - An episode is one salary point from the timeline (one row in long format)",
                "   - n_cao_year: Number of unique CAOs with salary episodes in that year",
                "   - n_jobgroups_year: Number of unique job groups with salary episodes in that year",
                "",
                "4. CAO-level structure distribution (section='cao_structure_distribution'):",
                "   - For each CAO, counts of rows, jobgroups, steps, worker_types, age_groups, educations",
                "   - Across CAOs: mean, median, min, max, p25, p75 for each metric",
                "",
                "5. FT hours development (section='ft_hours_by_year'):",
                "   - Statistics for full-time weekly hours by contract start year",
                "   - Shows how full-time hours evolve over time",
                "",
                "6. Salary monthly band (section='salary_monthly_band'):",
                "   - Event-level counts for normalized monthly amounts vs NL statutory min (Jan 1; band uses SALARY_ANALYSIS_MONTHLY_FLOOR_RELATIVE_MIN × floor) and cap",
                "   - n_increase_events, n_conversion_ok, n_band_eligible, n_dropped_* , share_* (denom: conversion_ok)",
                "   - See outputs/analysis/salary_monthly_band_summary.csv for the same row"
            ],
            "03_salary_slots_coverage": [
                "NOTES:",
                "This sheet summarizes the use of salary slots (timeline points) on the analytic wide sample.",
                "",
                "salary_index: Slot index k from columns salary_k_* (detected dynamically from the CSV header)",
                "n_rows_with_amount: Rows with a strictly positive, parseable salary amount in this slot",
                "  (amount <= 0, missing, or non-numeric counts as no amount)",
                "share_rows_with_amount: Proportion of all rows with a positive amount in this slot",
                "n_rows_with_start_date: Number of rows with a start date in this slot",
                "most_common_unit: Most frequently used pay unit for this slot (e.g., 'monthly', '4-week')",
                "share_most_common_unit: Proportion of rows using the most common unit",
                "",
                "Note: Wide rows can carry many timeline slots (salary_1_*, salary_2_*, ...); the count",
                "depends on extraction output, not a fixed cap."
            ],
            "04_worker_profile": [
                "NOTES:",
                "This sheet provides frequency distributions for worker-related dimensions (analytic wide sample).",
                "",
                "variable: The dimension being analyzed (worker_type, permanency, is_entry, age_group, education)",
                "category: The category value",
                "n_rows: Number of salary rows with this category",
                "share_rows: Proportion of all rows with this category",
                "",
                "For is_entry:",
                "  - True: Entry/aanloop scales (starting scales for new workers)",
                "  - False: Standard scales (not explicitly entry scales)"
            ],
            "05_amounts_by_unit": [
                "NOTES:",
                "This sheet analyzes salary amounts by pay unit to understand typical magnitudes.",
                "",
                "block: 'by_unit_and_index' (by unit and slot) or 'by_unit' (aggregated across all slots)",
                "salary_unit: Pay period unit (e.g., 'monthly', '4-week', 'weekly', 'hourly', 'annual')",
                "salary_index: Timeline slot index k, or NaN for aggregated statistics",
                "n: Number of salary episodes with this unit (and slot, if applicable); long table excludes amount <= 0",
                "mean, median, p25, p75, min, max: Descriptive statistics for salary amounts",
                "",
                "This helps understand what typical salary levels are for different units and",
                "to infer what ambiguous units like 'period' or 'unit' likely represent."
            ],
            "06_increase_percent": [
                "NOTES:",
                "This sheet analyzes general wage increase percentages specified in CAOs.",
                "",
                "block: 'overall' (all episodes) or 'by_year' (grouped by salary start year)",
                "",
                "For overall block:",
                "  - statistic: n, mean, median, p25, p75, min, max",
                "  - value: The statistic value",
                "",
                "For by_year block:",
                "  - statistic: year, n_year, mean_year, median_year",
                "  - value: The statistic value",
                "  - Only includes years with at least 3 non-missing increase percentages",
                "",
                "Note: increase_percent represents general percentage increases for wage tables",
                "as explicitly stated in the CAO (e.g., 3.00 for a +3% wage rise)."
            ],
            "07_level_trends_by_year": [
                "NOTES:",
                "This sheet shows salary level development over time for the dominant pay unit.",
                "",
                "block: 'all_episodes' (all salary episodes) or 'entry_scales' (entry scales only)",
                "salary_start_year: Year when the salary amount becomes effective",
                "primary_unit: The most common pay unit in the dataset (e.g., 'monthly')",
                "",
                "For all_episodes block:",
                "  - n_episodes: Number of salary episodes in primary_unit starting in that year",
                "  - mean_amount, median_amount: Average and median salary amounts",
                "  - p25_amount, p75_amount: 25th and 75th percentiles",
                "",
                "For entry_scales block:",
                "  - n_entry_episodes: Number of entry scale episodes in primary_unit",
                "  - mean_entry_amount, median_entry_amount: Average and median entry scale amounts",
                "",
                "Note: Only includes episodes with the primary_unit to ensure comparability.",
                "Long-format input excludes non-positive salary amounts (same rule as n_salary_points_per_row)."
            ],
            "08_row_slot_counts": [
                "NOTES:",
                "This sheet describes how many salary points (slots) each row carries (analytic sample; buckets with n=0 may be omitted or rare).",
                "",
                "block: 'overall' (all rows) or 'by_year' (grouped by contract start year)",
                "n_salary_points_per_row: Bucket — integers 0..10 count exact tallies; '11+' is rows with >= 11",
                "  positive salary amounts (coerced numerically; <= 0 does not count)",
                "n_rows: Number of rows in that bucket",
                "share_rows: Proportion of rows in that bucket",
                "",
                "For by_year block:",
                "  - contract_start_year: Year when the contract starts",
                "  - Same bucketing within each year",
                "",
                "Note: Maximum slot index in the file is determined by salary_k_* columns, not fixed."
            ],
            "10_increase_diff_only": [
                "NOTES:",
                "Event-level derived increases: consecutive within-row percentage changes of normalized monthly amounts.",
                "",
                "Rows with failed unit conversion or non-positive amounts are excluded from diff chains;",
                "conversion issues appear in salary_increase_conversion_diagnostics.csv.",
                "analysis_monthly_band_ok: amounts must lie between SALARY_ANALYSIS_MONTHLY_FLOOR_RELATIVE_MIN × NL statutory monthly min (salary_start_date)",
                "and SALARY_ANALYSIS_MONTHLY_CAP_EUR; otherwise increase_diff_only is NaN and merged pref is masked.",
                "Band exclusions are counted in 02_sample_overview (section salary_monthly_band) and salary_monthly_band_summary.csv.",
                "is_first_salary_in_file: 1 for every event on the earliest salary_start_date in that file (cao+file)."
            ],
            "11_increase_merge_vs_csv": [
                "NOTES:",
                "Comparison where both increase_csv_only and increase_diff_only are present.",
                "",
                "abs_diff: absolute difference (percentage points); within_0_1pp if abs_diff <= 0.1;",
                "abs_diff_gt_0_1 flags larger disagreements; sign_disagreement compares signs.",
                "",
                "Full comparable rows are also exported to outputs/analysis/salary_increase_csv_vs_diff_comparison.csv (sep=;)."
            ],
            "09_no_salary_analysis": [
                "NOTES:",
                "This sheet analyzes the full extract (not restricted to the analytic sample), including rows with n_salary_points_per_row == 0.",
                "",
                "Sections:",
                "",
                "1. Overall summary (block='overall_summary'):",
                "   - total_unique_files: Total number of unique files in the dataset",
                "   - files_with_salary_data: Number of files that contain salary information",
                "   - files_with_no_salary_data: Number of files with no salary information",
                "   - share: Proportion of total files",
                "",
                "2. By CAO (block='by_cao'):",
                "   - caos_with_only_no_salary_files: Number of CAOs that have ONLY files with no salary data",
                "   - caos_with_at_least_some_salary_files: Number of CAOs that have at least one file with salary data",
                "   - caos_with_at_least_one_no_salary_file: Number of CAOs that have at least one file with no salary (may also have salary files)",
                "   - total_files_in_caos_with_only_no_salary: Total number of files in CAOs that have no salary data at all",
                "",
                "3. By contract start year (block='by_contract_year'):",
                "   - n_files: Number of unique files with contracts starting in that year",
                "   - n_rows_no_salary: Number of rows with no salary data in that year",
                "   - n_rows_total: Total number of rows in that year",
                "   - share_rows_no_salary: Proportion of rows with no salary data",
                "   - n_cao: Number of unique CAOs with contracts starting in that year",
                "",
                "4. List of files with no salary data (block='no_salary_files_list'):",
                "   - Complete list of all files (cao_number + file_name) that have no salary data",
                "   - Includes metadata (id, TTW, dates, sector, sbi_code) if available"
            ]
        }
        
        # Add notes to sheets
        for sheet_name, note_lines in notes.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Find the last row with data
                max_row = ws.max_row
                # Start notes 2 rows below the data
                note_start_row = max_row + 3
                
                for i, note_line in enumerate(note_lines):
                    cell = ws.cell(row=note_start_row + i, column=1)
                    cell.value = note_line
                    # Make notes italic
                    try:
                        existing_font = cell.font
                        cell.font = Font(italic=True, name=existing_font.name if existing_font else 'Calibri', 
                                       size=existing_font.size if existing_font else 11, 
                                       bold=existing_font.bold if existing_font else False, 
                                       color=existing_font.color if existing_font and existing_font.color else None)
                    except Exception:
                        # Fallback if font access fails
                        cell.font = Font(italic=True)
        
        wb.save(OUTPUT_EXCEL)
        print(f"\n✓ Successfully created Excel workbook with {len(sheets)} sheets and explanatory notes")
        print(f"  Output: {OUTPUT_EXCEL}")
    except Exception as e:
        print(f"\n  ERROR: Could not write Excel file: {e}")
        import traceback
        traceback.print_exc()
        return
    
    print("\n" + "="*80)
    print("Script completed successfully!")
    print("="*80)


if __name__ == "__main__":
    main()

